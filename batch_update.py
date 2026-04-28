import openpyxl, shutil, os, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

AREA = 5.0
base = r"c:\Users\user\My Drive\KAIST MASc 2021\Laboratory Work\Protocol\overpotential calculation\For paper SI"

# Load all source workbooks once
wb_iv_main  = openpyxl.load_workbook(base + r"\IV All data Main.xlsx",  data_only=True)
wb_iv_other = openpyxl.load_workbook(base + r"\IV All data Other.xlsx", data_only=True)
wb_hfr      = openpyxl.load_workbook(base + r"\HFR+Proton All.xlsx",    data_only=True)

# (sample, iv_workbook, hfr_sheet, iv_i_col_BOL, hfr_val_col_BOL)
# iv_i_col_BOL : column of I for BOL; V = I+1; 30K offset +2; 75K offset +4
# hfr_val_col_BOL: column of HFR/Rp value for BOL; 30K +2; 75K +4
SAMPLES = [
    ("CN BM",     wb_iv_main,  "Main",  1,  2),
    ("CN Polyol", wb_iv_other, "Other", 1,  2),
    ("KB BM",     wb_iv_main,  "Main",  7,  8),
    ("KB Polyol", wb_iv_other, "Other", 7,  8),
    ("VC BM",     wb_iv_other, "Other", 13, 14),
    ("VC Polyol", wb_iv_main,  "Main",  13, 14),
    ("AB BM",     wb_iv_other, "Other", 19, 20),
    ("AB Polyol", wb_iv_main,  "Main",  19, 20),
]

DUR_OFFSET = {"BOL": 0, "30K": 2, "75K": 4}

# (cond_label, iv_sheet, dur_subfolder_map, hfr_bp_row)
#   hfr_bp_row: 2 = 0 bar BP, 3 = 1.5 bar BP
CONDITIONS = [
    ("O2",     "O2 0 BP",   {"BOL":"BOL","30K":"30K","75K":"75K"}, 2),
    ("O2 BP",  "O2 1.5 BP", {"BOL":"BOL","30K":"30K","75K":"75k"}, 3),  # 75k lowercase
    ("Air",    "Air 0 BP",  {"BOL":"BOL","30K":"30K","75K":"75K"}, 2),
    ("Air BP", "Air 1.5 BP",{"BOL":"BOL","30K":"30K","75K":"75K"}, 3),
]

# Exact target file names per (cond, sample)
TARGET_NAMES = {
    ("O2",    "CN BM"):     "CN BM o2.xlsx",
    ("O2",    "CN Polyol"): "CN Polyol O2.xlsx",
    ("O2",    "KB BM"):     "KB BM O2.xlsx",
    ("O2",    "KB Polyol"): "KB Polyol O2.xlsx",
    ("O2",    "VC BM"):     "VC BM O2.xlsx",
    ("O2",    "VC Polyol"): "VC Polyol O2.xlsx",
    ("O2",    "AB BM"):     "AB BM  O2.xlsx",   # double space
    ("O2",    "AB Polyol"): "AB Polyol O2.xlsx",

    ("O2 BP", "CN BM"):     "CN BM 1.5bp O2.xlsx",
    ("O2 BP", "CN Polyol"): "CN Polyol 1.5bp O2.xlsx",
    ("O2 BP", "KB BM"):     "KB BM 1.5bp O2.xlsx",
    ("O2 BP", "KB Polyol"): "KB Polyol 1.5bp O2.xlsx",
    ("O2 BP", "VC BM"):     "VC BM 1.5bp O2.xlsx",
    ("O2 BP", "VC Polyol"): "VC Polyol 1.5bp O2.xlsx",
    ("O2 BP", "AB BM"):     "AB BM 1.5bp O2.xlsx",
    ("O2 BP", "AB Polyol"): "AB Polyol 1.5bp O2.xlsx",

    ("Air",   "CN BM"):     "CN BM air.xlsx",
    ("Air",   "CN Polyol"): "CN Polyol air.xlsx",
    ("Air",   "KB BM"):     "KB BM air.xlsx",
    ("Air",   "KB Polyol"): "KB Polyol air.xlsx",
    ("Air",   "VC BM"):     "VC BM air.xlsx",
    ("Air",   "VC Polyol"): "VC Polyol air.xlsx",
    ("Air",   "AB BM"):     "AB BM Air.xlsx",   # capital A
    ("Air",   "AB Polyol"): "AB Polyol air.xlsx",

    ("Air BP","CN BM"):     "CN BM 1.5bp air.xlsx",
    ("Air BP","CN Polyol"): "CN Polyol 1.5bp air.xlsx",
    ("Air BP","KB BM"):     "KB BM 1.5bp air.xlsx",
    ("Air BP","KB Polyol"): "KB Polyol 1.5bp air.xlsx",
    ("Air BP","VC BM"):     "VC BM 1.5bp air.xlsx",
    ("Air BP","VC Polyol"): "VC Polyol 1.5bp air.xlsx",
    ("Air BP","AB BM"):     "AB BM 1.5bp air.xlsx",
    ("Air BP","AB Polyol"): "AB Polyol 1.5bp air.xlsx",
}

results = []

for cond, iv_sheet, dur_folders, hfr_bp_row in CONDITIONS:
    for dur in ["BOL","30K","75K"]:
        off = DUR_OFFSET[dur]
        dur_folder = dur_folders[dur]

        for samp, wb_iv, hfr_sheet, iv_i_base, hfr_col_base in SAMPLES:
            # --- Source I-V ---
            ws_iv = wb_iv[iv_sheet]
            ic = iv_i_base + off
            vc = iv_i_base + 1 + off
            src_iv = [(ws_iv.cell(r,ic).value, ws_iv.cell(r,vc).value)
                      for r in range(2, ws_iv.max_row+1)
                      if ws_iv.cell(r,ic).value is not None]

            # --- Source HFR / Rp ---
            ws_h = wb_hfr[hfr_sheet]
            hc = hfr_col_base + off
            src_hfr = ws_h.cell(hfr_bp_row, hc).value  # row 2 or 3
            src_rp  = ws_h.cell(4, hc).value             # row 4 always

            # --- Target path ---
            fname = TARGET_NAMES[(cond, samp)]
            tgt_dir  = os.path.join(base, "Overpotnital", cond, dur_folder)
            tgt_path = os.path.join(tgt_dir, fname)

            if not os.path.exists(tgt_path):
                results.append((cond, dur, samp, "FILE MISSING", "-", "-", "-", "-"))
                continue

            # Skip lock files (Excel has file open)
            lock = os.path.join(tgt_dir, "~$" + fname)
            if os.path.exists(lock):
                results.append((cond, dur, samp, "LOCKED (close in Excel)", "-", "-", "-", "-"))
                continue

            # --- Read target ---
            wb_t = openpyxl.load_workbook(tgt_path, data_only=True)
            ws_t = wb_t.active
            tgt_iv = [(ws_t.cell(r,1).value, ws_t.cell(r,2).value)
                      for r in range(4, ws_t.max_row+1)
                      if ws_t.cell(r,1).value is not None]
            tgt_hfr = ws_t.cell(4,4).value
            tgt_rp  = ws_t.cell(4,5).value
            tgt_rows = len(tgt_iv)
            wb_t.close()

            # --- Compare ---
            iv_match = (len(src_iv) == len(tgt_iv) and
                        all(abs(si*AREA/1000 - ti) <= 1e-8 and abs(sv - tv) <= 1e-8
                            for (si,sv),(ti,tv) in zip(src_iv, tgt_iv)))
            hfr_match = (src_hfr is not None and tgt_hfr is not None and
                         abs(src_hfr - tgt_hfr) <= 1e-10)
            rp_match  = (src_rp  is not None and tgt_rp  is not None and
                         abs(src_rp  - tgt_rp)  <= 1e-10)
            all_match = iv_match and hfr_match and rp_match

            stem    = fname.replace(".xlsx","")
            out_name = stem + ("_unchanged" if all_match else "_edited") + ".xlsx"
            out_path = os.path.join(tgt_dir, out_name)

            try:
                shutil.copy2(tgt_path, out_path)
            except PermissionError:
                results.append((cond, dur, samp, "LOCKED (close in Excel)", "-", "-", "-", "-"))
                continue

            if not all_match:
                wb_e = openpyxl.load_workbook(out_path)
                ws_e = wb_e.active

                # Write source rows
                for i, (si, sv) in enumerate(src_iv):
                    r = 4 + i
                    ws_e.cell(r,1).value = round(si*AREA/1000, 12)
                    ws_e.cell(r,2).value = sv
                    ws_e.cell(r,4).value = src_hfr
                    ws_e.cell(r,5).value = src_rp

                # Clear any extra rows if target had more rows than source
                for i in range(len(src_iv), tgt_rows):
                    r = 4 + i
                    for c in (1,2,4,5):
                        ws_e.cell(r,c).value = None

                wb_e.save(out_path)
                wb_e.close()

            changes = []
            if not iv_match:
                changes.append(f"IV(src={len(src_iv)},tgt={tgt_rows})")
            if not hfr_match:
                changes.append(f"HFR {tgt_hfr}→{src_hfr}")
            if not rp_match:
                changes.append(f"Rp {tgt_rp}→{src_rp}")

            status = "UNCHANGED" if all_match else ("EDITED: " + "; ".join(changes))
            results.append((cond, dur, samp, status, len(src_iv), tgt_rows,
                            src_hfr, src_rp))

wb_iv_main.close()
wb_iv_other.close()
wb_hfr.close()

# --- Print summary ---
print(f"{'Cond':<8} {'Dur':<4} {'Sample':<12} {'SrcR':>5} {'TgtR':>5}  Status")
print("="*90)
for row in results:
    cond, dur, samp, status, sr, tr, hfr, rp = row
    print(f"{cond:<8} {dur:<4} {samp:<12} {str(sr):>5} {str(tr):>5}  {status}")

unchanged = sum(1 for r in results if r[3]=="UNCHANGED")
edited    = sum(1 for r in results if r[3].startswith("EDITED"))
missing   = sum(1 for r in results if r[3]=="FILE MISSING")
print(f"\nTotal: {len(results)}  |  Unchanged: {unchanged}  |  Edited: {edited}  |  Missing: {missing}")
