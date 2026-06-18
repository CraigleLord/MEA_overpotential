"""
Overpot_grid_Main and Overpot_grid_Other with red markers at 300 mA/cm²
(instead of the 500 mA/cm² used in the originals).
Format identical to plot_overpotential_all_fixed.py build_bol_grid.
Output: Figure Plots (Claude AI)/Overpot_grid_*_300mA.png
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE     = os.path.dirname(os.path.abspath(__file__))
ALT_DATA = os.path.join(BASE, "Durability I-V figure", "tafel plot alt",
                         "Overpotential DATA raw alt")
OUT_DIR          = os.path.join(BASE, "Figure Plots (Claude AI)")
OUT_DIR_260603   = os.path.join(BASE, "260603 New data set (reproducibility)")
FINAL_PLOTS_DIR  = os.path.join(OUT_DIR_260603, "Final Plots")
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(OUT_DIR_260603, exist_ok=True)
os.makedirs(FINAL_PLOTS_DIR, exist_ok=True)

MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

BOL_CONDITIONS = [
    ("Air",    "BOL", "Air 0 BP"),
    ("Air BP", "BOL", "Air 1.5 BP"),
    ("O2",     "BOL", r"O$_2$ 0 BP"),
    ("O2 BP",  "BOL", r"O$_2$ 1.5 BP"),
]

FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

O2BP_STEMS = {
    "CN BM":     "CN BM 1.5bp O2",
    "KB BM":     "KB BM 1.5bp O2",
    "VC Polyol": "VC Polyol 1.5bp O2",
    "AB Polyol": "AB Polyol 1.5bp O2",
    "CN Polyol": "CN Polyol 1.5bp O2",
    "KB Polyol": "KB Polyol 1.5bp O2",
    "VC BM":     "VC BM 1.5bp O2",
    "AB BM":     "AB BM 1.5bp O2",
}

C_KIN  = "#FFA07A"
C_OHM  = "#90C090"
C_PRO  = "#9090D8"
C_MASS = "#F0E860"

MARKER_I = 500.0   # mA/cm²

# EIS-derived proton resistance overrides (replaces per-point Excel column D)
EIS_RPRO = {
    "CN BM": 0.020229,
}

# 260603 reproducibility data overrides: (cond_folder, stem) -> folder containing
# {dur_folder}/{stem}_edited.xlsx (dur_folder = "BOL"/"30K"/"75K")
REPRO_DATA_260603 = os.path.join(BASE, "260603 New data set (reproducibility)",
                                  "Overpotential DATA raw alt")
DATA_OVERRIDES = {
    ("Air BP", "KB BM 1.5bp air"): os.path.join(REPRO_DATA_260603, "Air BP", "KB BM"),
}


def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))
    if len(rows) < 5:
        raise ValueError("Too few data rows")
    wb.close()
    I_mA  = np.array([r[0] * 200 for r in rows])
    V     = np.array([r[1]        for r in rows])
    R_ohm = np.array([r[2]        for r in rows])
    R_pro = np.array([r[3]        for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def read_kin_params(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    wb.close()
    if not isinstance(intercept, (int, float)) or not isinstance(slope, (int, float)):
        raise ValueError(f"Non-numeric kin params: intercept={intercept!r}, slope={slope!r}")
    return float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    override = DATA_OVERRIDES.get((cond_folder, stem))
    if override:
        for dur_try in (dur_folder, dur_folder.lower()):
            for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
                p = os.path.join(override, dur_try, stem + sfx)
                if os.path.exists(p):
                    return p
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None


def draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
               fs=18, lw=1.6, show_legend=False, ymax=1.18):

    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    I_A  = I_mA / 200
    ln_j = np.log(I_mA / 1000)

    eta_kin = -(intercept + slope * ln_j)
    eta_ohm = R_ohm * I_A
    eta_pro = R_pro * I_A

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm_ = V_kin - eta_ohm
    V_pro_ = V_ohm_ - eta_pro

    V_kin_c  = np.maximum(V_kin,  V)
    V_ohm_c  = np.maximum(V_ohm_, V)
    V_pro_c  = np.maximum(V_pro_, V)

    I_plot = I_mA / 1000  # convert to A cm⁻² for display

    ax.fill_between(I_plot, V_kin_c,  V_erev,   color=C_KIN,  alpha=0.85, label="Kinetic")
    ax.fill_between(I_plot, V_ohm_c,  V_kin_c,  color=C_OHM,  alpha=0.85, label="Ohmic")
    ax.fill_between(I_plot, V_pro_c,  V_ohm_c,  color=C_PRO,  alpha=0.85, label="Proton")
    ax.fill_between(I_plot, V,        V_pro_c,  color=C_MASS, alpha=0.85, label="Mass Transport")

    ax.plot(I_plot, V, color="black", linewidth=lw, zorder=5)

    # Red markers at 0.2 and 1.0 A cm⁻² — label shows cell voltage (V) to 2 d.p.
    for j_tgt in (0.2, 1.0):           # A cm⁻²
        i_tgt = j_tgt * 1000           # mA cm⁻²
        if I_mA.min() <= i_tgt <= I_mA.max():
            v_tgt = float(np.interp(i_tgt, I_mA, V))
            ax.plot(j_tgt, v_tgt, marker="o", ms=6, color="red",
                    linestyle="none", zorder=10)
            ax.text(j_tgt + 0.08, v_tgt, f"{v_tgt:.2f}",
                    color="red", fontsize=fs, va="center", ha="left")

    # If the curve never reaches 1.0 A cm⁻², mark the end (highest current) point.
    # Red dot labelled with cell voltage; current density shown separately in blue
    # with an arrow pointing to the dot.
    j_max = I_mA[-1] / 1000            # peak (= last) point, A cm⁻²
    if j_max < 1.0:
        ax.plot(j_max, V[-1], marker="o", ms=6, color="red",
                linestyle="none", zorder=10)
        ax.text(j_max + 0.08, V[-1], f"{V[-1]:.2f}",
                color="red", fontsize=fs, va="center", ha="left")
        ax.annotate(f"{j_max:.2f} Acm$^{{-2}}$", xy=(j_max, V[-1]),
                    xytext=(j_max + 0.35, V[-1] + 0.30),
                    color="#1565C0", fontsize=fs, ha="left", va="bottom",
                    arrowprops=dict(arrowstyle="->", color="#1565C0", lw=1.2))

    ax.set_xlim(0, 3.5)
    ax.set_ylim(0.2, ymax)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    ax.xaxis.set_major_locator(plt.MultipleLocator(1))
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}"))
    ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)

    if show_legend:
        patches = [
            mpatches.Patch(color=C_KIN,  label="Kinetic"),
            mpatches.Patch(color=C_OHM,  label="Ohmic"),
            mpatches.Patch(color=C_PRO,  label="Proton"),
            mpatches.Patch(color=C_MASS, label="Mass Transport"),
        ]
        ax.legend(handles=patches, loc="upper right", fontsize=12,
                  frameon=False, facecolor="none", edgecolor="black",
                  handlelength=1.0, borderpad=0.3, labelspacing=0.15)


def build_bol_grid(sample_list, group_name, dur="BOL", show_titles=True, fname=None):
    nrows     = len(BOL_CONDITIONS)   # conditions → rows
    ncols     = len(sample_list)      # samples    → columns
    fs        = 18
    title_pad = fs * 0.67
    ymax      = 1.18

    ROW_LABEL_X = 0.03
    YLABEL_X    = 0.07
    LEFT        = 0.13

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(4.0 * ncols, 3.5 * nrows),
        dpi=300,
    )
    fig.subplots_adjust(left=LEFT, right=0.97, hspace=0.25, wspace=0.25)
    if show_titles:
        fig.supxlabel(r"Current Density (Acm$^{-2}$)", fontsize=fs, y=0.06)
        fig.supylabel("Cell Voltage (V)", fontsize=fs, x=YLABEL_X)

    # Pre-compute O2 BP kinetics per sample (intrinsic) at the requested duration
    kin_params = {}
    for samp in sample_list:
        o2bp_fp = find_file("O2 BP", dur, O2BP_STEMS[samp])
        if o2bp_fp:
            try:
                kin_params[samp] = read_kin_params(o2bp_fp)
            except Exception as e:
                print(f"  WARN: O2 BP kin read failed for {samp}: {e}")
                kin_params[samp] = (None, None)
        else:
            print(f"  WARN: O2 BP {dur} file not found for {samp}")
            kin_params[samp] = (None, None)

    for row_idx, (cond_folder, _dur_unused, cond_label) in enumerate(BOL_CONDITIONS):
        for col_idx, samp in enumerate(sample_list):
            ax         = axes[row_idx, col_idx]
            is_topleft = (row_idx == 0 and col_idx == 0)
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur, stem) if stem else None

            if fp is None:
                ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)
                ax.set_xlim(0, 3.5); ax.set_ylim(0.2, ymax)
            else:
                try:
                    I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                    kin_intercept, kin_slope = kin_params[samp]
                    if kin_intercept is not None:
                        intercept, slope = kin_intercept, kin_slope
                    if samp in EIS_RPRO:
                        R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                    erev = calc_erev(T_K, P_H2, P_O2)
                    draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                               fs=fs, lw=1.6,
                               show_legend=is_topleft,
                               ymax=ymax)
                except Exception as e:
                    print(f"  ERROR {samp}/{cond_folder}: {e}")
                    ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                            ha="center", va="center", fontsize=fs)

            if show_titles:
                # Row labels = condition names (left column only)
                if col_idx == 0:
                    pos   = axes[row_idx, 0].get_position()
                    y_fig = (pos.y0 + pos.y1) / 2
                    fig.text(ROW_LABEL_X, y_fig, cond_label,
                             fontsize=fs, fontweight="bold",
                             ha="center", va="center", rotation=90, clip_on=False)

    if fname is None:
        fname = f"Overpot_grid_{group_name}_300mA.png"
    for d in (OUT_DIR, OUT_DIR_260603, FINAL_PLOTS_DIR):
        out = os.path.join(d, fname)
        fig.savefig(out, dpi=300, bbox_inches="tight")
        print(f"Saved: {out}")
    plt.close(fig)


build_bol_grid(MAIN_SAMPLES,  "Main",  show_titles=False)
build_bol_grid(OTHER_SAMPLES, "Other", show_titles=False)

# 75K-only variant: no titles, no axis titles
build_bol_grid(MAIN_SAMPLES,  "Main",  dur="75K", show_titles=False,
               fname="Overpot_grid_Main_75K.png")
build_bol_grid(OTHER_SAMPLES, "Other", dur="75K", show_titles=False,
               fname="Overpot_grid_Other_75K.png")

print("Done.")
