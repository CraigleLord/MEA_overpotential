"""
Overpot_grid_{Main,Other}_300mA.png with an added 5th column: a per-sample bar
chart of mass-transport overpotential (eta_MT) across the four BOL conditions
(Air 0 BP, Air 1.5 BP, O2 0 BP, O2 1.5 BP), evaluated at a fixed current
density.

Two versions are produced:
  - eta_MT @ 0.5 A cm^-2  -> overwrites Overpot_grid_{Main,Other}_300mA.png
  - eta_MT @ 1.0 A cm^-2  -> Overpot_grid_{Main,Other}_300mA_MT1p0.png

Both written to (Figure Plots (Claude AI), 260603 root, Final Plots).
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE     = os.path.dirname(os.path.abspath(__file__))
ALT_DATA = os.path.join(BASE, "Durability I-V figure", "tafel plot alt",
                         "Overpotential DATA raw alt")
OUT_DIR        = os.path.join(BASE, "Figure Plots (Claude AI)")
OUT_DIR_260603 = os.path.join(BASE, "260603 New data set (reproducibility)")
FINAL_PLOTS_DIR = os.path.join(OUT_DIR_260603, "Final Plots")

MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

O2BP_STEMS = {
    "CN BM":     "CN BM 1.5bp O2",
    "KB BM":     "KB BM 1.5bp O2",
    "VC Polyol": "VC Polyol 1.5bp O2",
    "AB Polyol": "AB Polyol 1.5bp O2",
    "CN Polyol": "CN Polyol 1.5bp O2",
    "KB Polyol": "KB Polyol 1.5bp O2",
    "VC BM":     "VC BM 1.5bp O2",
    "AB BM":     "AB BM 1.5bp O2",
}

BOL_CONDITIONS = [
    ("Air",    "BOL", "Air 0 BP"),
    ("Air BP", "BOL", "Air 1.5 BP"),
    ("O2",     "BOL", r"O$_2$ 0 BP"),
    ("O2 BP",  "BOL", r"O$_2$ 1.5 BP"),
]

MT_TICK_LABELS = ["Air\n0bp", "Air\n1.5bp", "O$_2$\n0bp", "O$_2$\n1.5bp"]

C_KIN  = "#FFA07A"
C_OHM  = "#90C090"
C_PRO  = "#9090D8"
C_MASS = "#F0E860"

EIS_RPRO = {
    "CN BM": 0.020229,
}

# 260603 reproducibility data overrides: (cond_folder, stem) -> folder containing
# {dur_folder}/{stem}_edited.xlsx (dur_folder = "BOL"/"30K"/"75K")
REPRO_DATA_260603 = os.path.join(BASE, "260603 New data set (reproducibility)",
                                  "Overpotential DATA raw alt")
DATA_OVERRIDES = {
    ("Air BP", "KB BM 1.5bp air"):    os.path.join(REPRO_DATA_260603, "Air BP", "KB BM"),
    ("O2 BP",  "CN BM 1.5bp O2"):     os.path.join(REPRO_DATA_260603, "O2 BP", "CN BM"),
    ("O2 BP",  "KB BM 1.5bp O2"):     os.path.join(REPRO_DATA_260603, "O2 BP", "KB BM"),
    ("O2 BP",  "VC Polyol 1.5bp O2"): os.path.join(REPRO_DATA_260603, "O2 BP", "VC Polyol"),
    ("O2 BP",  "AB Polyol 1.5bp O2"): os.path.join(REPRO_DATA_260603, "O2 BP", "AB Polyol"),
}

# VC Polyol "final" dataset — every VC Polyol stem is routed here (takes precedence
# over the defaults above). Same {cond}/{dur}/Edited/{stem} layout as ALT_DATA.
VC_FINAL_BASE = os.path.join(BASE, "Overpotential DATA raw final")


def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))
    if len(rows) < 5:
        raise ValueError("Too few data rows")
    wb.close()
    I_mA  = np.array([r[0] * 200 for r in rows])
    V     = np.array([r[1]        for r in rows])
    R_ohm = np.array([r[2]        for r in rows])
    R_pro = np.array([r[3]        for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def read_kin_params(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    wb.close()
    if not isinstance(intercept, (int, float)) or not isinstance(slope, (int, float)):
        raise ValueError(f"Non-numeric kin params: intercept={intercept!r}, slope={slope!r}")
    return float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    if stem.startswith("VC Polyol"):   # final VC Polyol dataset takes precedence
        for dur_try in (dur_folder, dur_folder.lower()):
            folder = os.path.join(VC_FINAL_BASE, cond_folder, dur_try, "Edited")
            for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
                p = os.path.join(folder, stem + sfx)
                if os.path.exists(p):
                    return p
    override = DATA_OVERRIDES.get((cond_folder, stem))
    if override:
        for dur_try in (dur_folder, dur_folder.lower()):
            for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
                p = os.path.join(override, dur_try, stem + sfx)
                if os.path.exists(p):
                    return p
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None


def draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
               fs=18, lw=1.6, show_legend=False, ymax=1.18,
               legend_frameon=False, marker_I=500.0):

    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    I_A  = I_mA / 200
    ln_j = np.log(I_mA / 1000)

    eta_kin = -(intercept + slope * ln_j)
    eta_ohm = R_ohm * I_A
    eta_pro = R_pro * I_A

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm_ = V_kin - eta_ohm
    V_pro_ = V_ohm_ - eta_pro

    V_kin_c  = np.maximum(V_kin,  V)
    V_ohm_c  = np.maximum(V_ohm_, V)
    V_pro_c  = np.maximum(V_pro_, V)

    I_plot = I_mA / 1000  # A cm^-2

    ax.fill_between(I_plot, V_kin_c,  V_erev,   color=C_KIN,  alpha=0.85, label="Kinetic")
    ax.fill_between(I_plot, V_ohm_c,  V_kin_c,  color=C_OHM,  alpha=0.85, label="Ohmic")
    ax.fill_between(I_plot, V_pro_c,  V_ohm_c,  color=C_PRO,  alpha=0.85, label="Proton")
    ax.fill_between(I_plot, V,        V_pro_c,  color=C_MASS, alpha=0.85, label="Mass Transport")

    ax.plot(I_plot, V, color="black", linewidth=lw, zorder=5)

    i_m      = marker_I
    v_meas_m = float(np.interp(i_m, I_mA, V))
    ax.plot(i_m / 1000, v_meas_m, marker="o", ms=6, color="red", linestyle="none", zorder=10)
    ax.text(i_m / 1000 + 0.1, v_meas_m, f"{v_meas_m:.2f}",
            color="red", fontsize=fs, va="center", ha="left")

    ax.set_xlim(0, 3.5)
    ax.set_ylim(0.2, ymax)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    ax.xaxis.set_major_locator(plt.MultipleLocator(1))
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}"))
    ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)

    if show_legend:
        patches = [
            mpatches.Patch(color=C_KIN,  label="Kinetic"),
            mpatches.Patch(color=C_OHM,  label="Ohmic"),
            mpatches.Patch(color=C_PRO,  label="Proton"),
            mpatches.Patch(color=C_MASS, label="Mass Transport"),
        ]
        ax.legend(handles=patches, loc="upper right", fontsize=12,
                  frameon=legend_frameon, facecolor="none", edgecolor="black",
                  handlelength=1.0, borderpad=0.3, labelspacing=0.15)


def calc_eta_mt(I_mA, V, R_ohm, R_pro, erev, intercept, slope, target_I_mA,
                clamp_to_max=False):
    """Mass-transport overpotential (V) at target_I_mA mA/cm^2.

    If target_I_mA exceeds the measured current range: return NaN, unless
    clamp_to_max=True, in which case evaluate at the highest measured
    current instead.
    """
    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    if target_I_mA > I_mA.max():
        if not clamp_to_max:
            return np.nan
        target_I_mA = I_mA.max()

    i      = target_I_mA
    I_A_i  = i / 200
    ln_j_i = np.log(i / 1000)
    Ro_i   = float(np.interp(i, I_mA, R_ohm))
    Rp_i   = float(np.interp(i, I_mA, R_pro))
    V_i    = float(np.interp(i, I_mA, V))

    eta_kin = max(0.0, -(intercept + slope * ln_j_i))
    eta_ohm = max(0.0, Ro_i * I_A_i)
    eta_pro = max(0.0, Rp_i * I_A_i)
    return max(0.0, erev - eta_kin - eta_ohm - eta_pro - V_i)


def calc_eta_mt_at_voltage(I_mA, V, R_ohm, R_pro, erev, intercept, slope, v_target):
    """eta_MT at the current density that corresponds to v_target on the I-V curve."""
    idx_peak = int(np.argmax(I_mA))
    I_t  = I_mA[:idx_peak + 1]
    V_t  = V[:idx_peak + 1]
    sort_idx = np.argsort(V_t)
    V_s, I_s = V_t[sort_idx], I_t[sort_idx]
    if v_target < V_s[0] or v_target > V_s[-1]:
        return np.nan
    target_I = float(np.interp(v_target, V_s, I_s))
    return calc_eta_mt(I_mA, V, R_ohm, R_pro, erev, intercept, slope, target_I)


def calc_eta_kin_at_voltage(I_mA, V, R_ohm, R_pro, erev, intercept, slope, v_target):
    """Kinetic overpotential at the current density that corresponds to v_target on the I-V curve."""
    idx_peak = int(np.argmax(I_mA))
    I_t, V_t = I_mA[:idx_peak + 1], V[:idx_peak + 1]
    sort_idx  = np.argsort(V_t)
    V_s, I_s  = V_t[sort_idx], I_t[sort_idx]
    if v_target < V_s[0] or v_target > V_s[-1]:
        return np.nan
    j_mA  = float(np.interp(v_target, V_s, I_s))   # mA cm⁻²
    ln_j  = np.log(j_mA / 1000)                     # j in A cm⁻²
    return max(0.0, -(intercept + slope * ln_j))


def calc_eta_pro_at_voltage(I_mA, V, R_ohm, R_pro, erev, intercept, slope, v_target):
    """Proton-transfer overpotential at the current density corresponding to v_target."""
    idx_peak = int(np.argmax(I_mA))
    I_c, V_c, Rp_c = I_mA[:idx_peak + 1], V[:idx_peak + 1], R_pro[:idx_peak + 1]
    sort_idx = np.argsort(V_c)
    V_s, I_s = V_c[sort_idx], I_c[sort_idx]
    if v_target < V_s[0] or v_target > V_s[-1]:
        return np.nan
    j_mA = float(np.interp(v_target, V_s, I_s))
    Rp_i = float(np.interp(j_mA, I_c, Rp_c))
    return max(0.0, Rp_i * j_mA / 200)


def calc_components_at_voltage(I_mA, V, R_ohm, R_pro, erev, intercept, slope, v_target):
    """All four overpotential components (kin, ohm, pro, mt) in V at the current
    density where the I-V curve crosses v_target. Returns None if out of range."""
    idx_peak = int(np.argmax(I_mA))
    I_c, V_c, Ro_c, Rp_c = (I_mA[:idx_peak + 1], V[:idx_peak + 1],
                            R_ohm[:idx_peak + 1], R_pro[:idx_peak + 1])
    sort_idx = np.argsort(V_c)
    V_s, I_s = V_c[sort_idx], I_c[sort_idx]
    if v_target < V_s[0] or v_target > V_s[-1]:
        return None
    j_mA  = float(np.interp(v_target, V_s, I_s))
    I_A_i = j_mA / 200
    ln_j  = np.log(j_mA / 1000)
    Ro_i  = float(np.interp(j_mA, I_c, Ro_c))
    Rp_i  = float(np.interp(j_mA, I_c, Rp_c))
    V_i   = float(np.interp(j_mA, I_c, V_c))
    eta_kin = max(0.0, -(intercept + slope * ln_j))
    eta_ohm = max(0.0, Ro_i * I_A_i)
    eta_pro = max(0.0, Rp_i * I_A_i)
    eta_mt  = max(0.0, erev - eta_kin - eta_ohm - eta_pro - V_i)
    return eta_kin, eta_ohm, eta_pro, eta_mt


def calc_components_at_current(I_mA, V, R_ohm, R_pro, erev, intercept, slope, target_mA):
    """All four overpotential components (kin, ohm, pro, mt) in V at target_mA
    mA/cm^2. If the I-V curve does not reach target_mA, evaluate at the end
    (highest current) point of the curve instead."""
    idx_peak = int(np.argmax(I_mA))
    I_c, V_c, Ro_c, Rp_c = (I_mA[:idx_peak + 1], V[:idx_peak + 1],
                            R_ohm[:idx_peak + 1], R_pro[:idx_peak + 1])
    j_mA = min(target_mA, float(I_c.max()))   # clamp to end point if curve falls short
    I_A_i = j_mA / 200
    ln_j  = np.log(j_mA / 1000)
    Ro_i  = float(np.interp(j_mA, I_c, Ro_c))
    Rp_i  = float(np.interp(j_mA, I_c, Rp_c))
    V_i   = float(np.interp(j_mA, I_c, V_c))
    eta_kin = max(0.0, -(intercept + slope * ln_j))
    eta_ohm = max(0.0, Ro_i * I_A_i)
    eta_pro = max(0.0, Rp_i * I_A_i)
    eta_mt  = max(0.0, erev - eta_kin - eta_ohm - eta_pro - V_i)
    return eta_kin, eta_ohm, eta_pro, eta_mt


def _nice_upper(v):
    if v <= 0:
        return 0.0
    padded = v * 1.25
    mag = 10 ** math.floor(math.log10(padded))
    for mult in (1.0, 2.0, 5.0):
        step = mult * mag
        candidate = math.ceil(padded / step) * step
        if candidate / step <= 5:
            return candidate
    return math.ceil(padded * 1000) / 1000


def _set_mt_yticks(ax, y_max):
    if y_max <= 0.02:
        step, dec = 0.005, 3
    elif y_max <= 0.05:
        step, dec = 0.01, 2
    elif y_max <= 0.10:
        step, dec = 0.02, 2
    elif y_max <= 0.25:
        step, dec = 0.05, 2
    elif y_max <= 0.5:
        step, dec = 0.1, 1
    else:
        step, dec = 0.2, 1
    ax.yaxis.set_major_locator(plt.MultipleLocator(step))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.{dec}f}"))


# ── Pass 1: read all data + compute eta_MT at 0.5 and 1.0 A/cm^2 ───────────
GROUPS = {"Main": MAIN_SAMPLES, "Other": OTHER_SAMPLES}
cache      = {}   # (group, samp, cond_idx) -> dict of draw_panel kwargs, or None
mt_values  = {}   # (group, samp, cond_idx) -> (eta_mt_05, eta_mt_10)

for group_name, sample_list in GROUPS.items():
    for samp in sample_list:
        o2bp_fp = find_file("O2 BP", "BOL", O2BP_STEMS[samp])
        kin_intercept, kin_slope = (None, None)
        if o2bp_fp:
            try:
                kin_intercept, kin_slope = read_kin_params(o2bp_fp)
            except Exception as e:
                print(f"  WARN: O2 BP kin read failed for {samp}: {e}")

        for cond_idx, (cond_folder, dur_folder, _) in enumerate(BOL_CONDITIONS):
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None
            key  = (group_name, samp, cond_idx)
            if fp is None:
                print(f"  WARN: no file for {samp}/{cond_folder}")
                cache[key]     = None
                mt_values[key] = (np.nan, np.nan)
                continue
            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                if kin_intercept is not None:
                    intercept, slope = kin_intercept, kin_slope
                if samp in EIS_RPRO:
                    R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                erev = calc_erev(T_K, P_H2, P_O2)
                cache[key] = dict(I_mA=I_mA, V=V, R_ohm=R_ohm, R_pro=R_pro,
                                  erev=erev, intercept=intercept, slope=slope)
                mt05 = calc_eta_mt(I_mA, V, R_ohm, R_pro, erev, intercept, slope, 500.0)
                mt10 = calc_eta_mt(I_mA, V, R_ohm, R_pro, erev, intercept, slope, 1000.0,
                                   clamp_to_max=True)
                if np.isnan(mt05):
                    print(f"  NOTE: {samp}/{cond_folder} max I = {I_mA.max():.0f} mA/cm^2 (mt05=nan)")
                if I_mA.max() < 1000.0:
                    print(f"  NOTE: {samp}/{cond_folder} max I = {I_mA.max():.0f} mA/cm^2 "
                          f"-> mt10 clamped to {mt10:.4f} V at max I")
                mt_values[key] = (mt05, mt10)
            except Exception as e:
                print(f"  ERROR {samp}/{cond_folder}: {e}")
                cache[key]     = None
                mt_values[key] = (np.nan, np.nan)

YMAX_05 = 0.15  # fixed per user request (0-0.15 V, 0.05/tick)
YMAX_10 = _nice_upper(max(v[1] for v in mt_values.values() if not np.isnan(v[1])))
print(f"Shared y-max: eta_MT@0.5A = {YMAX_05:.4f} V, eta_MT@1.0A = {YMAX_10:.4f} V")

for (group_name, samp, cond_idx), (mt05, mt10) in mt_values.items():
    if not np.isnan(mt05) and mt05 > YMAX_05:
        print(f"  WARN: {group_name}/{samp}/{BOL_CONDITIONS[cond_idx][2]} "
              f"mt05={mt05:.4f} V exceeds bar y-axis max {YMAX_05} V")


# ── Pass 1b: parallel 75K cache (O2 BP 75K kinetics, matching Overpot_grid_*_75K) ──
cache75 = {}   # (group, samp, cond_idx) -> dict of component kwargs, or None
for group_name, sample_list in GROUPS.items():
    for samp in sample_list:
        o2bp_fp = find_file("O2 BP", "75K", O2BP_STEMS[samp])
        kin_intercept, kin_slope = (None, None)
        if o2bp_fp:
            try:
                kin_intercept, kin_slope = read_kin_params(o2bp_fp)
            except Exception as e:
                print(f"  WARN: O2 BP 75K kin read failed for {samp}: {e}")
        for cond_idx, (cond_folder, _dur, _) in enumerate(BOL_CONDITIONS):
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, "75K", stem) if stem else None
            key  = (group_name, samp, cond_idx)
            if fp is None:
                cache75[key] = None
                continue
            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                if kin_intercept is not None:
                    intercept, slope = kin_intercept, kin_slope
                if samp in EIS_RPRO:
                    R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                erev = calc_erev(T_K, P_H2, P_O2)
                cache75[key] = dict(I_mA=I_mA, V=V, R_ohm=R_ohm, R_pro=R_pro,
                                    erev=erev, intercept=intercept, slope=slope)
            except Exception as e:
                print(f"  WARN: 75K read failed {samp}/{cond_folder}: {e}")
                cache75[key] = None


# ── Pass 1c: grid-resolution caches (match Overpot_grid_* exactly) ──────────
# The waterfall j@V / V@j bar charts must reproduce the red marker values in the
# Overpot grids. The grids resolve files with ONLY the KB BM Air BP override
# (no VC Polyol "final" routing, no O2 BP repro) — VC Polyol comes from ALT_DATA.
GRID_DATA_OVERRIDES = {
    ("Air BP", "KB BM 1.5bp air"): os.path.join(REPRO_DATA_260603, "Air BP", "KB BM"),
}


def find_file_grid(cond_folder, dur_folder, stem):
    override = GRID_DATA_OVERRIDES.get((cond_folder, stem))
    if override:
        for dur_try in (dur_folder, dur_folder.lower()):
            for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
                p = os.path.join(override, dur_try, stem + sfx)
                if os.path.exists(p):
                    return p
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None


cache_g, cache_g75 = {}, {}   # grid-resolution I-V caches for the waterfall charts
for group_name, sample_list in GROUPS.items():
    for samp in sample_list:
        for cond_idx, (cond_folder, _dur, _) in enumerate(BOL_CONDITIONS):
            stem = FILE_STEMS.get((cond_folder, samp))
            key  = (group_name, samp, cond_idx)
            for tgt_cache, dur in ((cache_g, "BOL"), (cache_g75, "75K")):
                fp = find_file_grid(cond_folder, dur, stem) if stem else None
                if fp is None:
                    tgt_cache[key] = None
                    continue
                try:
                    I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                    tgt_cache[key] = dict(I_mA=I_mA, V=V, R_ohm=R_ohm, R_pro=R_pro,
                                          erev=calc_erev(T_K, P_H2, P_O2),
                                          intercept=intercept, slope=slope)
                except Exception as e:
                    print(f"  WARN: grid-cache {samp}/{cond_folder}/{dur}: {e}")
                    tgt_cache[key] = None


# ── Pass 2: build figures ───────────────────────────────────────────────────
def build_mt_grid(group_name, sample_list, target_idx, target_mA, y_max, out_filename):
    ncols = len(BOL_CONDITIONS)
    nrows = len(sample_list)
    fs        = 18
    title_pad = fs * 0.67
    ymax_iv   = 1.18

    FIG_W_IV  = 4.0
    FIG_W_BAR = 4.0
    total_w   = FIG_W_IV * ncols + FIG_W_BAR
    _s        = (FIG_W_IV * ncols) / total_w
    ROW_LABEL_X = 0.03 * _s
    YLABEL_X    = 0.07 * _s
    LEFT        = 0.13 * _s

    fig = plt.figure(figsize=(total_w, 3.5 * nrows), dpi=300)
    gs  = fig.add_gridspec(nrows, ncols + 1,
                            width_ratios=[FIG_W_IV] * ncols + [FIG_W_BAR])
    fig.subplots_adjust(left=LEFT, right=0.97, hspace=0.25, wspace=0.30)

    axes     = np.array([[fig.add_subplot(gs[r, c]) for c in range(ncols)]
                          for r in range(nrows)])
    bar_axes = [fig.add_subplot(gs[r, ncols]) for r in range(nrows)]

    # Nudge the 5th (bar) column right by 60 px to add breathing room vs. column 4
    shift_frac = 60.0 / (fig.get_figwidth() * fig.dpi)
    for ax_bar in bar_axes:
        pos = ax_bar.get_position()
        ax_bar.set_position([pos.x0 + shift_frac, pos.y0, pos.width, pos.height])

    iv_xc = (axes[0, 0].get_position().x0 + axes[0, ncols - 1].get_position().x1) / 2
    fig.supxlabel(r"Current Density (Acm$^{-2}$)", fontsize=fs, y=0.06, x=iv_xc)
    fig.supylabel("Cell Voltage (V)", fontsize=fs, x=YLABEL_X)

    for row_idx, samp in enumerate(sample_list):
        for col_idx, (cond_folder, dur_folder, cond_label) in enumerate(BOL_CONDITIONS):
            ax         = axes[row_idx, col_idx]
            is_topleft = (row_idx == 0 and col_idx == 0)
            data = cache[(group_name, samp, col_idx)]
            if data is None:
                ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)
                ax.set_xlim(0, 3.5); ax.set_ylim(0.2, ymax_iv)
            else:
                draw_panel(ax, **data, fs=fs, lw=1.6,
                           show_legend=is_topleft, ymax=ymax_iv)

            if row_idx == 0:
                ax.set_title(cond_label, fontsize=fs, fontweight="bold", pad=title_pad)
            if col_idx == 0:
                pos   = axes[row_idx, 0].get_position()
                y_fig = (pos.y0 + pos.y1) / 2
                fig.text(ROW_LABEL_X, y_fig, LABELS[samp],
                         fontsize=fs, fontweight="bold",
                         ha="center", va="center", rotation=90, clip_on=False)

        # ── Mass-transport bar panel ────────────────────────────────────────
        ax_bar = bar_axes[row_idx]
        vals = [mt_values[(group_name, samp, c)][target_idx] for c in range(ncols)]
        vals = [0.0 if np.isnan(v) else v for v in vals]
        x = np.arange(ncols)
        ax_bar.bar(x, vals, width=0.55, facecolor=C_MASS,
                   edgecolor="black", linewidth=0.5)
        ax_bar.set_xlim(-0.6, ncols - 0.4)
        ax_bar.set_ylim(0, y_max)
        _set_mt_yticks(ax_bar, y_max)
        ax_bar.set_xticks(x)
        if row_idx == nrows - 1:
            ax_bar.set_xticklabels(MT_TICK_LABELS, fontsize=fs)
        else:
            ax_bar.tick_params(labelbottom=False)
        ax_bar.tick_params(direction="out", top=False, right=False,
                           labelsize=fs, width=1.2, length=5)
        for spine in ax_bar.spines.values():
            spine.set_linewidth(1.2)
        if row_idx == 0:
            ax_bar.set_title(rf"$\eta_{{MT}}$ @ {target_mA/1000:.1f} A cm$^{{-2}}$",
                             fontsize=fs, fontweight="bold", pad=title_pad)

    # single centred y-label for the bar column (freeze-and-centre)
    for ax_bar in bar_axes:
        ax_bar.set_ylabel("Mass Transport η (V)", fontsize=fs)
    fig.canvas.draw()
    renderer = fig.canvas.get_renderer()
    y_c    = (bar_axes[0].get_position().y1 + bar_axes[-1].get_position().y0) / 2
    lbl_bb = bar_axes[0].yaxis.label.get_window_extent(renderer)
    x_c    = (lbl_bb.x0 + lbl_bb.x1) / 2 / (fig.get_figwidth() * fig.dpi)
    for ax_bar in bar_axes:
        ax_bar.yaxis.label.set_visible(False)
    fig.text(x_c, y_c, "Mass Transport η (V)",
             ha="center", va="center", rotation=90, fontsize=fs)

    for d in (OUT_DIR, OUT_DIR_260603, FINAL_PLOTS_DIR):
        out = os.path.join(d, out_filename)
        fig.savefig(out, dpi=300, bbox_inches="tight")
        print(f"Saved: {out}")
    plt.close(fig)


# ── Standalone bar-chart-only column (no I-V panels) ────────────────────────
def build_mt_bar_only_column(group_name, sample_list, target_idx, target_mA, y_max, out_filename):
    fs    = 18
    ncols = len(BOL_CONDITIONS)
    nsamp = len(sample_list)

    fig = plt.figure(figsize=(4.0, 3.2 * nsamp), constrained_layout=True)
    gs  = fig.add_gridspec(nsamp, 1)
    axes = [fig.add_subplot(gs[r, 0]) for r in range(nsamp)]

    x = np.arange(ncols)
    for r, samp in enumerate(sample_list):
        ax = axes[r]
        vals = [mt_values[(group_name, samp, ci)][target_idx] for ci in range(ncols)]
        vals = [0.0 if np.isnan(v) else v for v in vals]
        for ci, v in enumerate(vals):
            if v > y_max:
                print(f"  WARN: {group_name}/{samp}/{BOL_CONDITIONS[ci][2]} "
                      f"mt={v:.4f} V exceeds bar y-axis max {y_max} V")

        ax.bar(x, vals, width=0.55, facecolor=C_MASS, edgecolor="black", linewidth=0.5)
        ax.set_xlim(-0.6, ncols - 0.4)
        ax.set_ylim(0, y_max)
        _set_mt_yticks(ax, y_max)
        ax.set_xticks(x)
        if r == nsamp - 1:
            ax.set_xticklabels(MT_TICK_LABELS, fontsize=fs)
        else:
            ax.tick_params(labelbottom=False)
        ax.tick_params(direction="out", top=False, right=False,
                        labelsize=fs, width=1.2, length=5)
        for spine in ax.spines.values():
            spine.set_linewidth(1.2)

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_mt_bar_cond_rows(group_name, sample_list, v_target, out_filename):
    """4 rows (conditions) × nsamp samples; eta_MT (yellow) + proton (purple, thin).
    Proton fits inside the 0–0.25 V scale so both share the same left axis."""
    fs     = 18
    nconds = len(BOL_CONDITIONS)
    nsamp  = len(sample_list)

    mt_matrix, pro_matrix = [], []
    for cond_idx in range(nconds):
        mt_row, pro_row = [], []
        for samp in sample_list:
            data = cache[(group_name, samp, cond_idx)]
            if data is None:
                mt_row.append(0.0); pro_row.append(0.0)
            else:
                m = calc_eta_mt_at_voltage(**data, v_target=v_target)
                p = calc_eta_pro_at_voltage(**data, v_target=v_target)
                mt_row.append(0.0 if np.isnan(m) else m)
                pro_row.append(0.0 if np.isnan(p) else p)
        mt_matrix.append(mt_row); pro_matrix.append(pro_row)

    # Keep 0–0.25 V unless MT or proton exceeds the limit.
    Y_MAX    = 0.25
    data_max = max([v for row in (mt_matrix + pro_matrix) for v in row] + [0.0])
    if data_max > Y_MAX:
        Y_MAX = _nice_upper(data_max)

    MAIN_W, PRO_W = 0.42, 0.18
    x = np.arange(nsamp)

    fig = plt.figure(figsize=(5.5, 4.0 * nconds), constrained_layout=True)
    gs   = fig.add_gridspec(nconds, 1)
    axes = [fig.add_subplot(gs[r, 0]) for r in range(nconds)]

    for ci, ax in enumerate(axes):
        ax.bar(x - 0.10, mt_matrix[ci], width=MAIN_W,
               facecolor=C_MASS, edgecolor="black", linewidth=0.5)
        ax.bar(x + 0.21, pro_matrix[ci], width=PRO_W,
               facecolor=C_PRO, edgecolor="black", linewidth=0.5)
        ax.set_xlim(-0.6, nsamp - 0.4)
        ax.set_ylim(0, Y_MAX)
        _set_mt_yticks(ax, Y_MAX)
        ax.set_xticks(x)
        ax.tick_params(labelbottom=False)
        ax.tick_params(direction="out", top=False, right=False,
                       labelsize=fs, width=1.2, length=5)
        for spine in ax.spines.values():
            spine.set_linewidth(1.2)
        ax.spines["top"].set_visible(True)   # top border on the plot box

        # Legend on the Air 0 BP panel (first row)
        if ci == 0:
            handles = [
                mpatches.Patch(facecolor=C_PRO,  edgecolor="black", label="Proton"),
                mpatches.Patch(facecolor=C_MASS, edgecolor="black", label="Mass Transport"),
            ]
            ax.legend(handles=handles, loc="upper right", fontsize=fs,
                      frameon=False, handlelength=1.2, labelspacing=0.3)

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_kin_bar_cond_rows(group_name, sample_list, v_target, out_filename):
    """4 rows (conditions) × nsamp bars (samples); kinetic η evaluated at v_target V."""
    C_KIN_BAR = "#FFA500"   # orange
    fs     = 18
    nconds = len(BOL_CONDITIONS)
    nsamp  = len(sample_list)

    vals_matrix = []
    for cond_idx in range(nconds):
        row = []
        for samp in sample_list:
            data = cache[(group_name, samp, cond_idx)]
            if data is None:
                row.append(0.0)
            else:
                v = calc_eta_kin_at_voltage(**data, v_target=v_target)
                row.append(0.0 if np.isnan(v) else v)
        vals_matrix.append(row)

    # Tight shared y-range that brackets the data so sample differences are
    # visible: floor/ceil to 0.05, step 0.05 (→ 4–5 ticks).
    all_nonzero = [v for row in vals_matrix for v in row if v > 0]
    STEP  = 0.05
    y_min = math.floor(min(all_nonzero) / STEP) * STEP
    y_max = math.ceil(max(all_nonzero) / STEP) * STEP

    fig = plt.figure(figsize=(5.5, 4.0 * nconds), constrained_layout=True)
    gs   = fig.add_gridspec(nconds, 1)
    axes = [fig.add_subplot(gs[r, 0]) for r in range(nconds)]

    x = np.arange(nsamp)

    for ci, ax in enumerate(axes):
        ax.bar(x, vals_matrix[ci], width=0.55,
               facecolor=C_KIN_BAR, edgecolor="black", linewidth=0.5)
        ax.set_xlim(-0.6, nsamp - 0.4)
        ax.set_ylim(y_min, y_max)
        ax.yaxis.set_major_locator(plt.MultipleLocator(STEP))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
        ax.set_xticks(x)
        ax.tick_params(labelbottom=False)
        ax.tick_params(direction="out", top=False, right=False,
                       labelsize=fs, width=1.2, length=5)
        for spine in ax.spines.values():
            spine.set_linewidth(1.2)

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


# Combined component figure: per sample, 4 grouped bars (kinetic, ohmic, proton,
# mass transport). Rows = conditions, columns = the two voltages (0.7 V / 0.4 V).
COMP_INFO = [
    (0, "Kinetic",        "#FFA500"),   # orange
    (1, "Ohmic",          C_OHM),       # green
    (2, "Proton",         C_PRO),       # purple
    (3, "Mass Transport", C_MASS),      # yellow
]
COMBINED_VOLTS = [(0.70, "0.7 V"), (0.40, "0.4 V")]


def _combined_ymax():
    """Shared y-max across both groups, both voltages, all components."""
    m = 0.0
    for gname, slist in GROUPS.items():
        for samp in slist:
            for ci in range(len(BOL_CONDITIONS)):
                d = cache[(gname, samp, ci)]
                if d is None:
                    continue
                for vt, _ in COMBINED_VOLTS:
                    comps = calc_components_at_voltage(**d, v_target=vt)
                    if comps is not None:
                        m = max(m, max(comps))
    return max(_nice_upper(m), 0.01)


def build_combined_component_grid(group_name, sample_list, out_filename, y_max,
                                  y_major=0.2, y_minor=0.05, annotate=False,
                                  targets=None, calc_fn=None, mark_shortfall=False,
                                  with_75k=False, data_src=None, show_legend=True):
    """Columns = `targets`, a list of (value, label). `calc_fn(d, value)` returns
    the four components. Defaults reproduce the voltage-based chart.
    with_75k=True draws, per sample, 8 bars: each component as a BOL (solid) and
    a 75K (hatched) bar (75K values from cache75).
    data_src selects the source for the 4-bar (non-with_75k) path: BOL `cache`
    by default, or `cache75` for a 75K-only figure."""
    if targets is None:
        targets = COMBINED_VOLTS
    if calc_fn is None:
        calc_fn = lambda d, tv: calc_components_at_voltage(**d, v_target=tv)
    src = data_src if data_src is not None else cache

    fs     = 18
    nconds = len(BOL_CONDITIONS)
    nsamp  = len(sample_list)
    nvolts = len(targets)
    offsets = (np.arange(4) - 1.5) * 0.19   # 4 bars per sample
    bar_w   = 0.18
    x       = np.arange(nsamp)

    fig, axes = plt.subplots(nconds, nvolts,
                             figsize=(4.6 * nvolts, 3.6 * nconds),
                             squeeze=False, constrained_layout=True)

    for ci in range(nconds):
        for vi, (vt, vlabel) in enumerate(targets):
            ax = axes[ci][vi]

            def _label(xi, val, fsz):
                if val <= 0:
                    return
                if val > y_max:   # clipped bar → label just inside top
                    ax.text(xi, y_max - 0.004, f"{val:.2f}", rotation=90,
                            fontsize=fsz, ha="center", va="top", color="black")
                else:
                    ax.text(xi, val + 0.004, f"{val:.2f}", rotation=90,
                            fontsize=fsz, ha="center", va="bottom", color="black")

            if with_75k:
                unit = 0.115
                bw   = 0.10
                for comp_idx, comp_label, color in COMP_INFO:
                    bol_off = (2 * comp_idx - 3.5) * unit
                    k75_off = (2 * comp_idx + 1 - 3.5) * unit
                    bol_vals, k75_vals = [], []
                    for samp in sample_list:
                        db = cache[(group_name, samp, ci)]
                        dk = cache75[(group_name, samp, ci)]
                        cb = calc_fn(db, vt) if db is not None else None
                        ck = calc_fn(dk, vt) if dk is not None else None
                        bol_vals.append(0.0 if cb is None else cb[comp_idx])
                        k75_vals.append(0.0 if ck is None else ck[comp_idx])
                    ax.bar(x + bol_off, bol_vals, width=bw, facecolor=color,
                           edgecolor="black", linewidth=0.5, label=comp_label)
                    ax.bar(x + k75_off, k75_vals, width=bw, facecolor=color,
                           edgecolor="black", linewidth=0.5, hatch="//")
                    if annotate:
                        for xi, val in zip(x + bol_off, bol_vals):
                            _label(xi, val, 9)
                        for xi, val in zip(x + k75_off, k75_vals):
                            _label(xi, val, 9)
            else:
                for comp_idx, comp_label, color in COMP_INFO:
                    vals = []
                    for samp in sample_list:
                        d = src[(group_name, samp, ci)]
                        if d is None:
                            vals.append(0.0); continue
                        comps = calc_fn(d, vt)
                        vals.append(0.0 if comps is None else comps[comp_idx])
                    ax.bar(x + offsets[comp_idx], vals, width=bar_w,
                           facecolor=color, edgecolor="black", linewidth=0.5,
                           label=comp_label)
                    if annotate:
                        for xi, val in zip(x + offsets[comp_idx], vals):
                            _label(xi, val, 15)

            # Curves that fall short of the target current density: label the bar
            # group with the actual end-point current density (blue, matching the
            # grid colour scheme) and an arrow pointing to the group.
            if mark_shortfall:
                for xi, samp in zip(x, sample_list):
                    d = cache[(group_name, samp, ci)]
                    if d is None:
                        continue
                    cmax  = float(np.asarray(d["I_mA"]).max())   # end-point current, mA cm⁻²
                    j_end = cmax / 1000
                    if round(j_end, 2) < round(vt / 1000, 2):     # genuinely short of target
                        ax.annotate(f"{j_end:.2f}\nAcm$^{{-2}}$",
                                    xy=(xi, y_max * 0.58), xytext=(xi, y_max * 0.93),
                                    color="#1565C0", fontsize=12,
                                    ha="center", va="top",
                                    arrowprops=dict(arrowstyle="->",
                                                    color="#1565C0", lw=1.0))

            ax.set_xlim(-0.6, nsamp - 0.4)
            ax.set_ylim(0, y_max)
            ax.yaxis.set_major_locator(plt.MultipleLocator(y_major))
            ax.yaxis.set_minor_locator(plt.MultipleLocator(y_minor))
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
            ax.set_xticks(x)
            if ci == nconds - 1:
                ax.set_xticklabels([LABELS[s].replace("-", "\n") for s in sample_list],
                                   fontsize=fs)
            else:
                ax.tick_params(labelbottom=False)
            ax.tick_params(direction="out", top=False, right=False,
                           labelsize=fs, width=1.2, length=5)
            ax.tick_params(axis="y", which="minor", direction="out",
                           right=False, width=0.9, length=3)
            for spine in ax.spines.values():
                spine.set_linewidth(1.2)
            ax.spines["top"].set_visible(True)

            if ci == 0:
                ax.set_title(vlabel, fontsize=fs + 2, fontweight="bold")
            if vi == 0:
                ax.set_ylabel(BOL_CONDITIONS[ci][2], fontsize=fs, fontweight="bold")
            if ci == 0 and vi == 0 and show_legend:
                if with_75k:
                    comp_handles = [mpatches.Patch(facecolor=c, edgecolor="black", label=lbl)
                                    for _, lbl, c in COMP_INFO]
                    dur_handles = [
                        mpatches.Patch(facecolor="white", edgecolor="black", label="BOL"),
                        mpatches.Patch(facecolor="white", edgecolor="black", hatch="//", label="75K"),
                    ]
                    ax.legend(handles=comp_handles + dur_handles, fontsize=fs - 6,
                              frameon=False, ncol=3, handlelength=1.0,
                              labelspacing=0.3, columnspacing=0.8, loc="upper right")
                else:
                    ax.legend(fontsize=fs - 4, frameon=False, ncol=2,
                              handlelength=1.0, labelspacing=0.3, columnspacing=0.8,
                              loc="upper right")

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_stacked_dur_grid(group_name, sample_list, v_target, out_filename):
    """1 column × 4 rows (conditions). Per sample: two grouped stacked bars
    (BOL solid, 75K hatched), each stacking the 4 components at v_target.
    Values match Overpot_grid_*_300mA (BOL) and Overpot_grid_*_75K (75K)."""
    fs     = 18
    nconds = len(BOL_CONDITIONS)
    nsamp  = len(sample_list)
    x      = np.arange(nsamp)
    w      = 0.38
    DUR_SPEC = [("BOL", -0.21, "", cache), ("75K", +0.21, "//", cache75)]

    # components[ci][dur][si] -> (kin, ohm, pro, mt)
    components = {}
    y_top = 0.0
    for ci in range(nconds):
        components[ci] = {}
        for dur, _off, _h, src in DUR_SPEC:
            comps_row = []
            for samp in sample_list:
                d = src[(group_name, samp, ci)]
                c = calc_components_at_voltage(**d, v_target=v_target) if d is not None else None
                c = (0.0, 0.0, 0.0, 0.0) if c is None else c
                comps_row.append(c)
                y_top = max(y_top, sum(c))
            components[ci][dur] = comps_row

    y_max = max(math.ceil(y_top / 0.2) * 0.2, 0.2)

    fig = plt.figure(figsize=(6.0, 3.4 * nconds), constrained_layout=True)
    gs   = fig.add_gridspec(nconds, 1)
    axes = [fig.add_subplot(gs[r, 0]) for r in range(nconds)]

    for ci, ax in enumerate(axes):
        for dur, off, hatch, _src in DUR_SPEC:
            bottoms = np.zeros(nsamp)
            for comp_idx, comp_label, color in COMP_INFO:
                vals = np.array([components[ci][dur][si][comp_idx] for si in range(nsamp)])
                ax.bar(x + off, vals, width=w, bottom=bottoms,
                       facecolor=color, edgecolor="black", linewidth=0.5, hatch=hatch)
                bottoms += vals

        ax.set_xlim(-0.6, nsamp - 0.4)
        ax.set_ylim(0, y_max)
        _set_mt_yticks(ax, y_max)
        ax.set_xticks(x)
        if ci == nconds - 1:
            ax.set_xticklabels([LABELS[s].replace("-", "\n") for s in sample_list], fontsize=fs)
        else:
            ax.tick_params(labelbottom=False)
        ax.tick_params(direction="out", top=False, right=False,
                       labelsize=fs, width=1.2, length=5)
        for spine in ax.spines.values():
            spine.set_linewidth(1.2)
        ax.spines["top"].set_visible(True)
        ax.set_ylabel(BOL_CONDITIONS[ci][2], fontsize=fs, fontweight="bold")

        if ci == 0:
            comp_handles = [mpatches.Patch(facecolor=c, edgecolor="black", label=lbl)
                            for _, lbl, c in COMP_INFO]
            dur_handles = [
                mpatches.Patch(facecolor="white", edgecolor="black", label="BOL"),
                mpatches.Patch(facecolor="white", edgecolor="black", hatch="//", label="75K"),
            ]
            ax.legend(handles=comp_handles + dur_handles, fontsize=fs - 6, frameon=False,
                      ncol=3, handlelength=1.0, labelspacing=0.3, columnspacing=0.8,
                      loc="upper right")

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_voltage_waterfall_grid(group_name, sample_list, out_filename):
    """4 rows (conditions) × 2 cols (0.2 / 1.0 A cm^-2). Per sample a waterfall
    of CELL VOLTAGE: 75K (dark) at the bottom + BOL increment (light) on top.
    BOL value labelled inside the bar, 75K value on top. No reference markers."""
    fs       = 16
    nconds   = len(BOL_CONDITIONS)
    nsamp    = len(sample_list)
    currents = [(200.0, r"0.2 Acm$^{-2}$"), (1000.0, r"1.0 Acm$^{-2}$")]
    LIGHT, DARK = "#cccccc", "#555555"
    x = np.arange(nsamp)

    def v_at(d, tgt):
        if d is None:
            return np.nan
        I = np.asarray(d["I_mA"]); V = np.asarray(d["V"])
        ip = int(np.argmax(I))
        return float(np.interp(tgt, I[:ip + 1], V[:ip + 1]))   # clamps at curve ends

    data, vmax = {}, 0.0
    for ci in range(nconds):
        for tgt, _ in currents:
            for si, samp in enumerate(sample_list):
                vb = v_at(cache_g[(group_name, samp, ci)], tgt)
                vk = v_at(cache_g75[(group_name, samp, ci)], tgt)
                data[(ci, tgt, si)] = (vb, vk)
                if not np.isnan(vb):
                    vmax = max(vmax, vb)
    y_max = max(math.ceil(vmax / 0.2) * 0.2, 0.2)

    fig, axes = plt.subplots(nconds, 2, figsize=(5.2 * 2, 3.4 * nconds),
                             squeeze=False, constrained_layout=True)

    for ci in range(nconds):
        for col, (tgt, clabel) in enumerate(currents):
            ax = axes[ci][col]
            vb_arr = np.nan_to_num(np.array([data[(ci, tgt, si)][0] for si in range(nsamp)]))
            vk_arr = np.nan_to_num(np.array([data[(ci, tgt, si)][1] for si in range(nsamp)]))
            incr   = np.maximum(vb_arr - vk_arr, 0.0)

            ax.bar(x, vk_arr, width=0.55, facecolor=DARK, edgecolor="black",
                   linewidth=0.5, label="75K")
            ax.bar(x, incr, bottom=vk_arr, width=0.55, facecolor=LIGHT,
                   edgecolor="black", linewidth=0.5, label="BOL")

            for si in range(nsamp):
                vb, vk = vb_arr[si], vk_arr[si]
                if vb > 0:
                    ax.text(x[si], vk / 2, f"{vk:.2f}", ha="center", va="center",
                            fontsize=fs - 2, color="white")            # 75K inside
                    ax.text(x[si], vb + 0.01 * y_max, f"{vb:.2f}", ha="center",
                            va="bottom", fontsize=fs - 2, color="black")  # BOL on top

            ax.set_xlim(-0.6, nsamp - 0.4)
            ax.set_ylim(0, y_max)
            ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
            ax.set_xticks(x)
            if ci == nconds - 1:
                ax.set_xticklabels([LABELS[s].replace("-", "\n") for s in sample_list], fontsize=fs)
            else:
                ax.tick_params(labelbottom=False)
            ax.tick_params(direction="out", top=False, right=False,
                           labelsize=fs, width=1.2, length=5)
            for sp in ax.spines.values():
                sp.set_linewidth(1.2)
            ax.spines["top"].set_visible(True)

            if ci == 0:
                ax.set_title(clabel, fontsize=fs + 2, fontweight="bold")
            if col == 0:
                ax.set_ylabel(BOL_CONDITIONS[ci][2], fontsize=fs, fontweight="bold")
            if ci == 0 and col == 0:
                handles = [mpatches.Patch(facecolor=LIGHT, edgecolor="black", label="BOL"),
                           mpatches.Patch(facecolor=DARK, edgecolor="black", label="75K")]
                ax.legend(handles=handles, fontsize=fs - 2, frameon=False, loc="upper right")

    out = os.path.join(FINAL_PLOTS_DIR, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_waterfall_grid(group_name, sample_list, mode, out_subdir, out_filename,
                         style="waterfall"):
    """4 rows (conditions) × 2 cols. Per sample BOL & 75K.
      mode='j' : value = current density (A cm^-2) at 0.7 V and 0.4 V
      mode='v' : value = cell voltage (V) at 0.2 and 1.0 A cm^-2
      style='waterfall' : 75K dark bottom + BOL increment light on top
                          (75K labelled on top, BOL inside)
      style='grouped'   : BOL (light) and 75K (dark) as separate side-by-side
                          bars, each value labelled on top
    Saved into its own subfolder of Final Plots."""
    fs       = 16
    nconds   = len(BOL_CONDITIONS)
    nsamp    = len(sample_list)
    LIGHT, DARK = "#cccccc", "#555555"
    x = np.arange(nsamp)

    if mode == "j":
        cols = [(0.70, "0.7 V"), (0.40, "0.4 V")]
        def val(d, tgt):                       # current density (A cm^-2) at voltage tgt
            if d is None:
                return np.nan
            I = np.asarray(d["I_mA"]); V = np.asarray(d["V"])
            ip = int(np.argmax(I)); I, V = I[:ip + 1], V[:ip + 1]
            order = np.argsort(V); Vs, Is = V[order], I[order]
            if tgt < Vs[0] or tgt > Vs[-1]:
                return np.nan
            return float(np.interp(tgt, Vs, Is)) / 1000.0
        col_step = {0.70: 0.2, 0.40: 0.5}
    else:
        cols = [(200.0, r"0.2 Acm$^{-2}$"), (1000.0, r"1.0 Acm$^{-2}$")]
        def val(d, tgt):                       # cell voltage (V) at current tgt mA cm^-2
            if d is None:
                return np.nan
            I = np.asarray(d["I_mA"]); V = np.asarray(d["V"])
            ip = int(np.argmax(I))
            return float(np.interp(tgt, I[:ip + 1], V[:ip + 1]))
        col_step = {200.0: 0.2, 1000.0: 0.2}

    data = {}
    for ci in range(nconds):
        for tgt, _ in cols:
            for si, samp in enumerate(sample_list):
                vb = val(cache_g[(group_name, samp, ci)], tgt)
                vk = val(cache_g75[(group_name, samp, ci)], tgt)
                data[(ci, tgt, si)] = (vb, vk)

    def col_max(tgt):
        vals = [data[(ci, tgt, si)][0] for ci in range(nconds) for si in range(nsamp)]
        vals = [v for v in vals if not np.isnan(v)]
        return max(vals) if vals else 0.0

    if mode == "j":                            # per-column scale (magnitudes differ)
        col_ymax = {t: max(math.ceil(col_max(t) * 1.15 / col_step[t]) * col_step[t], col_step[t])
                    for t, _ in cols}
    else:                                      # shared scale for voltage
        gm = max(col_max(t) for t, _ in cols) * 1.12
        ymax_all = max(math.ceil(gm / 0.2) * 0.2, 0.2)
        col_ymax = {t: ymax_all for t, _ in cols}

    fig, axes = plt.subplots(nconds, 2, figsize=(5.2 * 2, 3.4 * nconds),
                             squeeze=False, constrained_layout=True)

    for ci in range(nconds):
        for col, (tgt, clabel) in enumerate(cols):
            ax    = axes[ci][col]
            y_max = col_ymax[tgt]
            vb_arr = np.nan_to_num(np.array([data[(ci, tgt, si)][0] for si in range(nsamp)]))
            vk_arr = np.nan_to_num(np.array([data[(ci, tgt, si)][1] for si in range(nsamp)]))

            if style == "grouped":
                gw = 0.34
                ax.bar(x - 0.18, vb_arr, width=gw, facecolor=LIGHT, edgecolor="black",
                       linewidth=0.5, label="BOL")
                ax.bar(x + 0.18, vk_arr, width=gw, facecolor=DARK, edgecolor="black",
                       linewidth=0.5, label="75K")
                for si in range(nsamp):
                    if vb_arr[si] > 0:
                        ax.text(x[si] - 0.18, vb_arr[si] + 0.012 * y_max, f"{vb_arr[si]:.2f}",
                                ha="center", va="bottom", fontsize=fs - 4, color="black")
                    if vk_arr[si] > 0:
                        ax.text(x[si] + 0.18, vk_arr[si] + 0.012 * y_max, f"{vk_arr[si]:.2f}",
                                ha="center", va="bottom", fontsize=fs - 4, color="black")
            elif style in ("bol", "k75"):
                vals  = vb_arr if style == "bol" else vk_arr
                color = LIGHT  if style == "bol" else DARK
                lbl   = "BOL"  if style == "bol" else "75K"
                ax.bar(x, vals, width=0.55, facecolor=color, edgecolor="black",
                       linewidth=0.5, label=lbl)
                for si in range(nsamp):
                    if vals[si] > 0:
                        ax.text(x[si], vals[si] + 0.012 * y_max, f"{vals[si]:.2f}",
                                ha="center", va="bottom", fontsize=fs - 3, color="black")
            else:
                incr = np.maximum(vb_arr - vk_arr, 0.0)
                ax.bar(x, vk_arr, width=0.55, facecolor=DARK, edgecolor="black",
                       linewidth=0.5, label="75K")
                ax.bar(x, incr, bottom=vk_arr, width=0.55, facecolor=LIGHT,
                       edgecolor="black", linewidth=0.5, label="BOL")
                for si in range(nsamp):
                    vb, vk = vb_arr[si], vk_arr[si]
                    if vb > 0:
                        ax.text(x[si], vk / 2, f"{vk:.2f}", ha="center", va="center",
                                fontsize=fs - 3, color="white")               # 75K inside
                        ax.text(x[si], vb + 0.012 * y_max, f"{vb:.2f}", ha="center",
                                va="bottom", fontsize=fs - 3, color="black")  # BOL on top

            ax.set_xlim(-0.6, nsamp - 0.4)
            ax.set_ylim(0, y_max)
            ax.yaxis.set_major_locator(plt.MultipleLocator(col_step[tgt]))
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
            ax.set_xticks(x)
            if ci == nconds - 1:
                ax.set_xticklabels([LABELS[s].replace("-", "\n") for s in sample_list], fontsize=fs)
            else:
                ax.tick_params(labelbottom=False)
            ax.tick_params(direction="out", top=False, right=False,
                           labelsize=fs, width=1.2, length=5)
            for sp in ax.spines.values():
                sp.set_linewidth(1.2)
            ax.spines["top"].set_visible(True)

            if ci == 0:
                ax.set_title(clabel, fontsize=fs + 2, fontweight="bold")
            if col == 0:
                ax.set_ylabel(BOL_CONDITIONS[ci][2], fontsize=fs, fontweight="bold")
            if ci == 0 and col == 0:
                if style == "bol":
                    handles = [mpatches.Patch(facecolor=LIGHT, edgecolor="black", label="BOL")]
                elif style == "k75":
                    handles = [mpatches.Patch(facecolor=DARK, edgecolor="black", label="75K")]
                else:
                    handles = [mpatches.Patch(facecolor=LIGHT, edgecolor="black", label="BOL"),
                               mpatches.Patch(facecolor=DARK, edgecolor="black", label="75K")]
                ax.legend(handles=handles, fontsize=fs - 2, frameon=False, loc="upper right")

    out_folder = os.path.join(FINAL_PLOTS_DIR, out_subdir)
    os.makedirs(out_folder, exist_ok=True)
    out = os.path.join(out_folder, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


def build_75k_jV_grid(group_name, sample_list, out_folder_abs, out_filename):
    """4 rows (conditions) × 2 cols (0.7 V / 0.4 V). 75K-only current-density
    bars (dark) with value labels on top. Same y-scales as the Main_barchart panels."""
    fs       = 16
    nconds   = len(BOL_CONDITIONS)
    nsamp    = len(sample_list)
    cols     = [(0.70, "0.7 V"), (0.40, "0.4 V")]
    col_ymax = {0.70: 0.8, 0.40: 2.5}
    col_step = {0.70: 0.2, 0.40: 0.5}
    DARK     = "#555555"
    x = np.arange(nsamp)

    def jval(d, tgt):
        if d is None:
            return np.nan
        I = np.asarray(d["I_mA"]); V = np.asarray(d["V"])
        ip = int(np.argmax(I)); I, V = I[:ip + 1], V[:ip + 1]
        order = np.argsort(V); Vs, Is = V[order], I[order]
        if tgt < Vs[0] or tgt > Vs[-1]:
            return np.nan
        return float(np.interp(tgt, Vs, Is)) / 1000.0

    fig, axes = plt.subplots(nconds, 2, figsize=(5.2 * 2, 3.4 * nconds),
                             squeeze=False, constrained_layout=True)
    for ci in range(nconds):
        for col, (tgt, clabel) in enumerate(cols):
            ax    = axes[ci][col]
            y_max = col_ymax[tgt]
            vals  = np.nan_to_num(np.array(
                [jval(cache_g75[(group_name, samp, ci)], tgt) for samp in sample_list]))
            ax.bar(x, vals, width=0.55, facecolor=DARK, edgecolor="black", linewidth=0.5)
            for si in range(nsamp):
                if vals[si] > 0:
                    ax.text(x[si], vals[si] + 0.012 * y_max, f"{vals[si]:.2f}",
                            ha="center", va="bottom", fontsize=fs - 3, color="black")
            ax.set_xlim(-0.6, nsamp - 0.4)
            ax.set_ylim(0, y_max)
            ax.yaxis.set_major_locator(plt.MultipleLocator(col_step[tgt]))
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
            ax.set_xticks(x)
            if ci == nconds - 1:
                ax.set_xticklabels([LABELS[s].replace("-", "\n") for s in sample_list], fontsize=fs)
            else:
                ax.tick_params(labelbottom=False)
            ax.tick_params(direction="out", top=False, right=False,
                           labelsize=fs, width=1.2, length=5)
            for sp in ax.spines.values():
                sp.set_linewidth(1.2)
            ax.spines["top"].set_visible(True)
            if ci == 0:
                ax.set_title(clabel, fontsize=fs + 2, fontweight="bold")
            if col == 0:
                ax.set_ylabel(BOL_CONDITIONS[ci][2], fontsize=fs, fontweight="bold")

    os.makedirs(out_folder_abs, exist_ok=True)
    out = os.path.join(out_folder_abs, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


build_mt_grid("Main",  MAIN_SAMPLES,  0, 500.0,  YMAX_05, "Overpot_grid_Main_300mA_MT0p5.png")
build_mt_grid("Other", OTHER_SAMPLES, 0, 500.0,  YMAX_05, "Overpot_grid_Other_300mA_MT0p5.png")
build_mt_grid("Main",  MAIN_SAMPLES,  1, 1000.0, YMAX_10, "Overpot_grid_Main_300mA_MT1p0.png")
build_mt_grid("Other", OTHER_SAMPLES, 1, 1000.0, YMAX_10, "Overpot_grid_Other_300mA_MT1p0.png")

build_mt_bar_cond_rows("Main",  MAIN_SAMPLES,  0.40, "Bar_MT_1p0_Main.png")
build_mt_bar_cond_rows("Other", OTHER_SAMPLES, 0.40, "Bar_MT_1p0_Other.png")

# Combined component figures: one per group; per sample 4 bars (components),
# rows = conditions, columns = voltages (0.7 V / 0.4 V). Shared y-axis 0–0.6 V.
# Full-range pair: y extended to 0.8 to leave headroom for on-top value labels
build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_Main.png",  0.7,
                              annotate=True, with_75k=True)
build_combined_component_grid("Other", OTHER_SAMPLES, "Bar_Components_Other.png", 0.7,
                              annotate=True, with_75k=True)

# 75K-only: 4 bars per sample (BOL removed)
build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_75K_Main.png", 0.7,
                              annotate=True, data_src=cache75)

# Zoomed pair: y 0–0.3, 0.1 per tick (kinetic bars clip but small components read clearly)
build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_0p3_Main.png",  0.3,
                              y_major=0.1, y_minor=0.05, annotate=True)
build_combined_component_grid("Other", OTHER_SAMPLES, "Bar_Components_0p3_Other.png", 0.3,
                              y_major=0.1, y_minor=0.05, annotate=True)

# Same component breakdown, but columns = two current densities instead of voltages:
# 0.2 A cm^-2 (high-voltage region) and 1.0 A cm^-2 (low-voltage region).
COMBINED_CURRENTS = [(200.0, r"0.2 Acm$^{-2}$"), (1000.0, r"1.0 Acm$^{-2}$")]
_calc_at_current = lambda d, tv: calc_components_at_current(**d, target_mA=tv)

build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_J_0p3_Main.png",  0.7,
                              y_major=0.2, y_minor=0.05, annotate=True,
                              targets=COMBINED_CURRENTS, calc_fn=_calc_at_current)
build_combined_component_grid("Other", OTHER_SAMPLES, "Bar_Components_J_0p3_Other.png", 0.7,
                              y_major=0.2, y_minor=0.05, annotate=True,
                              targets=COMBINED_CURRENTS, calc_fn=_calc_at_current)

# 8-bar (BOL + 75K) component figures, columns = 0.2 and 1.0 A cm^-2
build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_J_Main.png",  0.7,
                              annotate=True, with_75k=True,
                              targets=COMBINED_CURRENTS, calc_fn=_calc_at_current)
build_combined_component_grid("Other", OTHER_SAMPLES, "Bar_Components_J_Other.png", 0.7,
                              annotate=True, with_75k=True,
                              targets=COMBINED_CURRENTS, calc_fn=_calc_at_current)

# 75K-only: 4 bars per sample, columns = 0.2 and 1.0 A cm^-2
build_combined_component_grid("Main",  MAIN_SAMPLES,  "Bar_Components_J_75K_Main.png", 0.7,
                              annotate=True, data_src=cache75, show_legend=False,
                              targets=COMBINED_CURRENTS, calc_fn=_calc_at_current)

# Stacked BOL-vs-75K overpotential breakdown: 4 condition rows, grouped stacked
# bars per sample. One PNG per group × voltage (values match the Overpot grids).
build_stacked_dur_grid("Main",  MAIN_SAMPLES,  0.70, "Bar_Stack_BOL75K_0p7V_Main.png")
build_stacked_dur_grid("Other", OTHER_SAMPLES, 0.70, "Bar_Stack_BOL75K_0p7V_Other.png")
build_stacked_dur_grid("Main",  MAIN_SAMPLES,  0.40, "Bar_Stack_BOL75K_0p4V_Main.png")
build_stacked_dur_grid("Other", OTHER_SAMPLES, 0.40, "Bar_Stack_BOL75K_0p4V_Other.png")

# Cell-voltage waterfall (BOL/75K) at 0.2 and 1.0 A cm^-2
build_voltage_waterfall_grid("Main",  MAIN_SAMPLES,  "Bar_Vcell_J_Main.png")
build_voltage_waterfall_grid("Other", OTHER_SAMPLES, "Bar_Vcell_J_Other.png")

# 4 waterfall grids (2 cols × 4 condition rows), one per own folder:
#   current density @ 0.7/0.4 V  and  cell voltage @ 0.2/1.0 A cm^-2, Main & Other
build_waterfall_grid("Main",  MAIN_SAMPLES,  "j", "Waterfall Main current density",
                     "Bar_Waterfall_Main_jV.png")
build_waterfall_grid("Main",  MAIN_SAMPLES,  "v", "Waterfall Main cell voltage",
                     "Bar_Waterfall_Main_Vcell.png")
build_waterfall_grid("Other", OTHER_SAMPLES, "j", "Waterfall Other current density",
                     "Bar_Waterfall_Other_jV.png")
build_waterfall_grid("Other", OTHER_SAMPLES, "v", "Waterfall Other cell voltage",
                     "Bar_Waterfall_Other_Vcell.png")

# Separated (grouped BOL vs 75K) companion of each waterfall, same folders
build_waterfall_grid("Main",  MAIN_SAMPLES,  "j", "Waterfall Main current density",
                     "Bar_Grouped_Main_jV.png",    style="grouped")
build_waterfall_grid("Main",  MAIN_SAMPLES,  "v", "Waterfall Main cell voltage",
                     "Bar_Grouped_Main_Vcell.png",  style="grouped")
build_waterfall_grid("Other", OTHER_SAMPLES, "j", "Waterfall Other current density",
                     "Bar_Grouped_Other_jV.png",   style="grouped")
build_waterfall_grid("Other", OTHER_SAMPLES, "v", "Waterfall Other cell voltage",
                     "Bar_Grouped_Other_Vcell.png", style="grouped")

# BOL-only and 75K-only single-series companions, same folders
for _g, _samps in (("Main", MAIN_SAMPLES), ("Other", OTHER_SAMPLES)):
    for _mode, _tag, _sub in (("j", "jV", "current density"), ("v", "Vcell", "cell voltage")):
        _folder = f"Waterfall {_g} {_sub}"
        build_waterfall_grid(_g, _samps, _mode, _folder,
                             f"Bar_BOL_{_g}_{_tag}.png", style="bol")
        build_waterfall_grid(_g, _samps, _mode, _folder,
                             f"Bar_75K_{_g}_{_tag}.png", style="k75")

# 75K-only current-density grid (Main), saved into the Main_barchart folder
build_75k_jV_grid("Main", MAIN_SAMPLES,
                  os.path.join(FINAL_PLOTS_DIR, "Main_barchart"), "Bar_75K_jV_Main.png")

print("Done.")
