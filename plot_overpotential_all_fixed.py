"""
Fixed overpotential breakdown plots — all colored bands clipped to the measured I-V curve.
Generates:
  - Overpot_grid_Main/Other.png          (4-sample × 4-condition BOL grid)
  - Overpot_dur_*_Main/Other.png         (4-sample × 3-durability per condition)
Output: Durability I-V figure/tafel plot alt/overpotential decoupled/
Reads:  Durability I-V figure/tafel plot alt/Overpotential DATA raw alt/
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE    = os.path.dirname(os.path.abspath(__file__))
ALT_DIR = os.path.join(BASE, "Durability I-V figure", "tafel plot alt")
ALT_DATA = os.path.join(ALT_DIR, "Overpotential DATA raw alt")
OUT_DIR  = os.path.join(ALT_DIR, "overpotential decoupled")
os.makedirs(OUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Sample groups
# ---------------------------------------------------------------------------
MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

# Durability rows for the dur-grid figures
DURS = [("BOL", "BOL"), ("30K", "30K"), ("75K", "75K")]

# Conditions
ALL_CONDITIONS = [
    ("Air",    "Air_0bp",  "Air 0 BP",      False),
    ("Air BP", "Air_15bp", "Air 1.5 BP",    True),
    ("O2",     "O2_0bp",   r"O$_2$ 0 BP",   False),
    ("O2 BP",  "O2_15bp",  r"O$_2$ 1.5 BP", True),
]

# Conditions as used by the BOL-only grid (label text only)
BOL_CONDITIONS = [
    ("Air",    "BOL", "Air 0 BP"),
    ("Air BP", "BOL", "Air 1.5 BP"),
    ("O2",     "BOL", r"O$_2$ 0 BP"),
    ("O2 BP",  "BOL", r"O$_2$ 1.5 BP"),
]

FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------
def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))
    if len(rows) < 5:
        raise ValueError("Too few data rows")
    wb.close()
    I_mA  = np.array([r[0] * 200 for r in rows])
    V     = np.array([r[1]        for r in rows])
    R_ohm = np.array([r[2]        for r in rows])
    R_pro = np.array([r[3]        for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None

# ---------------------------------------------------------------------------
# Panel drawing — all bands clipped to measured I-V curve
# ---------------------------------------------------------------------------
C_KIN  = "#FFA07A"
C_OHM  = "#90C090"
C_PRO  = "#9090D8"
C_MASS = "#F0E860"

ANNOT = (
    "RH100\nPt 5 wt%\n"
    r"0.05 mg$_\mathregular{Pt}$/cm$^2$"
    "\nIC 0.8, N212"
)


def draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
               fs=21, lw=1.6, show_legend=False, show_annot=False,
               annot_str="", ymax=1.2, show_xlabel=False, show_ylabel=False):

    # Trim fold-back: only use data up to the peak current density
    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    I_A  = I_mA / 200
    ln_j = np.log(I_mA / 1000)

    eta_kin = -(intercept + slope * ln_j)
    eta_ohm = R_ohm * I_A
    eta_pro = R_pro * I_A

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm  = V_kin - eta_ohm
    V_pro  = V_ohm - eta_pro

    # Clip every boundary so no colored region extends below the measured curve
    V_kin_c = np.maximum(V_kin, V)
    V_ohm_c = np.maximum(V_ohm, V)
    V_pro_c = np.maximum(V_pro, V)

    ax.fill_between(I_mA, V_kin_c, V_erev,  color=C_KIN,  alpha=0.85, label="Kinetic")
    ax.fill_between(I_mA, V_ohm_c, V_kin_c, color=C_OHM,  alpha=0.85, label="Ohmic")
    ax.fill_between(I_mA, V_pro_c, V_ohm_c, color=C_PRO,  alpha=0.85, label="Proton")
    ax.fill_between(I_mA, V,       V_pro_c, color=C_MASS, alpha=0.85, label="Mass Transport")

    ax.plot(I_mA, V, color="black", linewidth=lw, zorder=5)

    i_m      = 500.0
    v_kin_m  = float(np.interp(i_m, I_mA, V_kin_c))
    v_meas_m = float(np.interp(i_m, I_mA, V))
    ax.plot(i_m, v_kin_m,  marker="o", ms=6, color="red", linestyle="none", zorder=10)
    ax.plot(i_m, v_meas_m, marker="o", ms=6, color="red", linestyle="none", zorder=10)

    # Separate labels if the two values are too close to read without overlap
    min_gap = 0.10          # minimum V between label centres (~1.2× font height)
    mid = (v_kin_m + v_meas_m) / 2
    if abs(v_kin_m - v_meas_m) < min_gap:
        v_kin_txt  = mid + min_gap / 2
        v_meas_txt = mid - min_gap / 2
    else:
        v_kin_txt  = v_kin_m
        v_meas_txt = v_meas_m

    ax.text(i_m + 100, v_kin_txt,  f"{v_kin_m:.2f}",
            color="red", fontsize=fs, va="center", ha="left")
    ax.text(i_m + 100, v_meas_txt, f"{v_meas_m:.2f}",
            color="red", fontsize=fs, va="center", ha="left")

    ax.set_xlim(0, 3500)
    ax.set_ylim(0.2, ymax)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    if show_xlabel:
        ax.set_xlabel(r"Current Density (mAcm$^{-2}$)", fontsize=fs)
    if show_ylabel:
        ax.set_ylabel("Cell Voltage (V)", fontsize=fs)
    ax.xaxis.set_major_locator(plt.MultipleLocator(1000))
    ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)

    if show_legend:
        patches = [
            mpatches.Patch(color=C_KIN,  label="Kinetic"),
            mpatches.Patch(color=C_OHM,  label="Ohmic"),
            mpatches.Patch(color=C_PRO,  label="Proton"),
            mpatches.Patch(color=C_MASS, label="Mass Transport"),
        ]
        ax.legend(handles=patches, loc="upper right", fontsize=10,
                  frameon=True, facecolor="white", edgecolor="black",
                  handlelength=1.0, borderpad=0.3, labelspacing=0.15)

    if show_annot and annot_str:
        ax.text(0.98, 0.97, annot_str,
                transform=ax.transAxes, ha="right", va="top",
                fontsize=fs, linespacing=1.4)


def _empty_panel(ax, fs, ymax):
    ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
            ha="center", va="center", fontsize=fs)
    ax.set_xlim(0, 3500)
    ax.set_ylim(0.2, ymax)


# ---------------------------------------------------------------------------
# 1.  BOL-only 4×4 grid  (rows=samples, cols=conditions)
# ---------------------------------------------------------------------------
def build_bol_grid(sample_list, group_name):
    ncols = len(BOL_CONDITIONS)
    nrows = len(sample_list)
    fs        = 18
    label_x   = -0.022 * fs   # scales with fs: -0.462 at fs=21, -0.396 at fs=18
    title_pad = fs * 0.67      # scales with fs: ~14 at fs=21, ~12 at fs=18

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(4.0 * ncols, 3.5 * nrows),
        dpi=300,
    )
    fig.subplots_adjust(hspace=0.35, wspace=0.35)

    for row_idx, samp in enumerate(sample_list):
        for col_idx, (cond_folder, dur_folder, cond_label) in enumerate(BOL_CONDITIONS):
            ax      = axes[row_idx, col_idx]
            ymax    = 1.18
            is_topleft = (row_idx == 0 and col_idx == 0)
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None

            if fp is None:
                _empty_panel(ax, fs, ymax); continue

            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                erev = calc_erev(T_K, P_H2, P_O2)
                draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                           fs=fs, lw=1.6,
                           show_legend=is_topleft,
                           show_annot=False,
                           annot_str=ANNOT, ymax=ymax,
                           show_xlabel=is_topleft, show_ylabel=is_topleft)
            except Exception as e:
                print(f"  ERROR {samp}/{cond_folder}: {e}")
                ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)

            if row_idx == 0:
                ax.set_title(cond_label, fontsize=fs, fontweight="bold", pad=title_pad)
            if col_idx == 0:
                ax.text(label_x, 0.5, LABELS[samp],
                        transform=ax.transAxes, fontsize=fs, fontweight="bold",
                        ha="center", va="center", rotation=90, clip_on=False)

    out = os.path.join(OUT_DIR, f"Overpot_grid_{group_name}.png")
    fig.savefig(out, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {out}")


# ---------------------------------------------------------------------------
# 2.  Durability grid  (rows=BOL/30K/75K, cols=samples) per condition
# ---------------------------------------------------------------------------
def build_dur_grid(sample_list, cond_folder, safe_name, is_15bp, group_name):
    nrows = len(DURS)
    ncols = len(sample_list)
    fs        = 18
    ymax      = 1.18
    label_x   = -0.022 * fs
    title_pad = fs * 0.67

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(4.0 * ncols, 3.5 * nrows),
        dpi=300,
    )
    fig.subplots_adjust(hspace=0.35, wspace=0.35)

    for row_idx, (dur_label, dur_folder) in enumerate(DURS):
        for col_idx, samp in enumerate(sample_list):
            ax         = axes[row_idx, col_idx]
            is_topleft = (row_idx == 0 and col_idx == 0)
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None

            if fp is None:
                _empty_panel(ax, fs, ymax)
            else:
                try:
                    I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                    erev = calc_erev(T_K, P_H2, P_O2)
                    draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                               fs=fs, lw=1.6,
                               show_legend=is_topleft,
                               ymax=ymax,
                               show_xlabel=is_topleft, show_ylabel=is_topleft)
                except Exception as e:
                    print(f"  ERROR {samp}/{dur_label}/{cond_folder}: {e}")
                    ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                            ha="center", va="center", fontsize=fs)

            if row_idx == 0:
                ax.set_title(LABELS[samp], fontsize=fs, fontweight="bold", pad=title_pad)
            if col_idx == 0:
                ax.text(label_x, 0.5, dur_label,
                        transform=ax.transAxes, fontsize=fs, fontweight="bold",
                        ha="center", va="center", rotation=90, clip_on=False)

    out = os.path.join(OUT_DIR, f"Overpot_dur_{safe_name}_{group_name}.png")
    fig.savefig(out, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {out}")


# ---------------------------------------------------------------------------
# Run all
# ---------------------------------------------------------------------------
build_bol_grid(MAIN_SAMPLES,  "Main")
build_bol_grid(OTHER_SAMPLES, "Other")

for cond_folder, safe_name, _, is_15bp in ALL_CONDITIONS:
    build_dur_grid(MAIN_SAMPLES,  cond_folder, safe_name, is_15bp, "Main")
    build_dur_grid(OTHER_SAMPLES, cond_folder, safe_name, is_15bp, "Other")

print("Done.")
