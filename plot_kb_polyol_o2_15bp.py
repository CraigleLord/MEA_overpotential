"""Plot I-V and power density for KB Polyol, O2 1.5 bar BP."""

from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import openpyxl

ROOT    = Path(__file__).parent
SRC     = ROOT / "LSV_comparison.xlsx"
OUT_DIR = ROOT / "LSV_reference_style_png"
OUT_DIR.mkdir(exist_ok=True)

COLOR = (120/255, 175/255, 220/255)   # powder blue (KB family)
# O2 1.5 bar BP is the 4th block: start col = 10 (cols J:K)
COL   = 10

def read_series(wb, col):
    ws = wb["KB_Polyol"]
    i_vals, v_vals = [], []
    for r in range(3, ws.max_row + 1):
        a = ws.cell(row=r, column=col).value
        b = ws.cell(row=r, column=col + 1).value
        if a is None or b is None:
            continue
        try:
            ai, bi = float(a), float(b)
        except (TypeError, ValueError):
            continue
        if ai < -1e9 or bi < -1e9:
            continue
        i_vals.append(ai)
        v_vals.append(bi)
    return np.asarray(i_vals), np.asarray(v_vals)

def main():
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.weight": "bold",
        "axes.labelweight": "bold",
        "axes.linewidth": 1.4,
        "xtick.direction": "in",
        "ytick.direction": "in",
        "xtick.major.width": 1.2,
        "ytick.major.width": 1.2,
        "xtick.major.size": 5,
        "ytick.major.size": 5,
    })

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    I, V = read_series(wb, COL)
    P = I * V
    idx = int(np.argmax(P))
    print(f"KB Polyol O2 1.5bar BP: n={P.size}, "
          f"P_max={P[idx]:.1f} mW/cm² at i={I[idx]:.1f}, V={V[idx]:.3f}")

    line_kw = dict(color=COLOR, linestyle=(0, (1, 1.6)), linewidth=2.2,
                   dash_capstyle="round")

    fig, (axL, axR) = plt.subplots(1, 2, figsize=(11.5, 4.4))
    fig.subplots_adjust(left=0.075, right=0.985, bottom=0.155, top=0.965, wspace=0.25)

    # LSV panel
    axL.plot(I, V, **line_kw)
    axL.plot([I[idx]], [V[idx]], marker="o", markersize=8,
             markerfacecolor=COLOR, markeredgecolor=COLOR,
             linestyle="None", zorder=5)

    # Power panel
    axR.plot(I, P, **line_kw)

    for ax in (axL, axR):
        ax.set_xlim(0, 3500)
        ax.set_xticks(np.arange(0, 3501, 500))
        ax.tick_params(axis="both", which="major", labelsize=11)
        for spine in ax.spines.values():
            spine.set_linewidth(1.4)
        ax.minorticks_off()
        ax.grid(False)

    axL.set_ylim(0.2, 1.2)
    axL.set_yticks(np.arange(0.2, 1.21, 0.2))
    axR.set_ylim(0, 1200)
    axR.set_yticks(np.arange(0, 1201, 200))

    axL.set_xlabel("Current Density (mAcm$^{-2}$)", fontsize=13, fontweight="bold")
    axL.set_ylabel("Cell Voltage (V)",              fontsize=13, fontweight="bold")
    axR.set_xlabel("Current Density (mAcm$^{-2}$)", fontsize=13, fontweight="bold")
    axR.set_ylabel("Power Density (mWcm$^{-2}$)",   fontsize=13, fontweight="bold")

    # Legend
    handle = Line2D([0], [0], color=COLOR, linestyle=(0, (1, 1.6)), linewidth=2.2,
                    label="KB Polyol")
    axL.legend(handles=[handle], loc="upper left", frameon=False,
               fontsize=10.5, handlelength=2.4, bbox_to_anchor=(0.005, 0.99))

    # Condition annotation
    annot = "\n".join([
        "H$_2$/O$_2$", "RH100", "Pt 5 wt%",
        r"0.05 mg$_{\mathrm{Pt}}$/cm$^{2}$",
        "IC 0.8, N212", "1.5 bar$_{g}$",
    ])
    axL.text(0.985, 0.985, annot, transform=axL.transAxes,
             ha="right", va="top", fontsize=10.5, fontweight="bold", linespacing=1.35)

    out = OUT_DIR / "KB_Polyol_O2_15bp_combined.png"
    fig.savefig(out, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
