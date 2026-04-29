"""
4×4 overpotential-breakdown grid figures (BOL only).
Rows = samples, Columns = conditions (Air 0BP | Air 1.5BP | O2 0BP | O2 1.5BP).
Produces: Overpotential Graphs/Overpot_grid_Main.png
          Overpotential Graphs/Overpot_grid_Other.png
Legend + condition note appear only in the top-left panel.
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Sample groups
# ---------------------------------------------------------------------------
MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

ROW_LABELS = {
    "CN BM":     "CN-BM",
    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol",
    "AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol",
    "KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",
    "AB BM":     "AB-BM",
}

# ---------------------------------------------------------------------------
# Conditions (columns) — BOL only
# ---------------------------------------------------------------------------
CONDITIONS = [
    ("Air",    "BOL", "Air 0 BP"),
    ("Air BP", "BOL", "Air 1.5 BP"),
    ("O2",     "BOL", r"O$_2$ 0 BP"),
    ("O2 BP",  "BOL", r"O$_2$ 1.5 BP"),
]

# ---------------------------------------------------------------------------
# File stems per (cond_folder, sample)
# ---------------------------------------------------------------------------
FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

# ---------------------------------------------------------------------------
# Physics helpers
# ---------------------------------------------------------------------------
def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]

    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value

    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))

    if len(rows) < 5:
        raise ValueError("Too few data rows")

    I_mA  = np.array([r[0] * 200   for r in rows])   # mA cm⁻²
    V     = np.array([r[1]          for r in rows])
    R_ohm = np.array([r[2]          for r in rows])
    R_pro = np.array([r[3]          for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    folder = os.path.join(BASE, "Overpotnital", cond_folder, dur_folder, "Edited")
    for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
        p = os.path.join(folder, stem + sfx)
        if os.path.exists(p):
            return p
    return None

# ---------------------------------------------------------------------------
# Draw one overpotential panel
# ---------------------------------------------------------------------------
C_KIN  = "#FFA07A"
C_OHM  = "#90C090"
C_PRO  = "#9090D8"
C_MASS = "#F0E860"

def draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
               fs=7, lw=0.8, show_legend=False, show_annot=False, annot_str="",
               ymax=1.2):

    I_A   = I_mA / 200
    ln_j  = np.log(I_mA / 1000)

    eta_kin  = -(intercept + slope * ln_j)
    eta_ohm  = R_ohm * I_A
    eta_pro  = R_pro * I_A

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm  = V_kin  - eta_ohm
    V_pro  = V_ohm  - eta_pro
    V_pro_clipped = np.maximum(V_pro, V)

    ax.fill_between(I_mA, V_kin,  V_erev,        color=C_KIN,  alpha=0.85,
                    label="Kinetic")
    ax.fill_between(I_mA, V_ohm,  V_kin,          color=C_OHM,  alpha=0.85,
                    label="Ohmic")
    ax.fill_between(I_mA, V_pro,  V_ohm,          color=C_PRO,  alpha=0.85,
                    label="Proton")
    ax.fill_between(I_mA, V,      V_pro_clipped,  color=C_MASS, alpha=0.85,
                    label="Mass Transport")

    ax.plot(I_mA, V, color="black", linewidth=lw, zorder=5)

    # Red dots at 500 mA cm⁻²
    i_m = 500.0
    v_kin_m  = float(np.interp(i_m, I_mA, V_kin))
    v_meas_m = float(np.interp(i_m, I_mA, V))
    ax.plot(i_m, v_kin_m,  marker="o", ms=3, color="red",
            linestyle="none", zorder=10)
    ax.plot(i_m, v_meas_m, marker="o", ms=3, color="red",
            linestyle="none", zorder=10)
    ax.text(i_m + 60, v_kin_m,  f"{v_kin_m:.2f}",
            color="red", fontsize=fs - 1, va="center", ha="left")
    ax.text(i_m + 60, v_meas_m, f"{v_meas_m:.2f}",
            color="red", fontsize=fs - 1, va="center", ha="left")

    ax.set_xlim(0, 3500)
    ax.set_ylim(0.2, ymax)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=0.6, length=3)
    ax.set_xlabel(r"Current Density (mAcm$^{-2}$)", fontsize=fs)
    ax.set_ylabel("Cell Voltage (V)", fontsize=fs)
    ax.xaxis.set_major_locator(plt.MultipleLocator(1000))
    ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
    for spine in ax.spines.values():
        spine.set_linewidth(0.6)

    if show_legend:
        patches = [
            mpatches.Patch(color=C_KIN,  label="Kinetic"),
            mpatches.Patch(color=C_OHM,  label="Ohmic"),
            mpatches.Patch(color=C_PRO,  label="Proton"),
            mpatches.Patch(color=C_MASS, label="Mass Transport"),
        ]
        ax.legend(handles=patches, loc="center right", fontsize=fs - 1,
                  frameon=True, facecolor="white", edgecolor="black",
                  handlelength=1.2, borderpad=0.4, labelspacing=0.2)

    if show_annot and annot_str:
        ax.text(0.98, 0.97, annot_str,
                transform=ax.transAxes, ha="right", va="top",
                fontsize=fs - 1, linespacing=1.4)

# ---------------------------------------------------------------------------
# Build one grid figure
# ---------------------------------------------------------------------------
ANNOT = (
    "RH100\nPt 5 wt%\n"
    r"0.05 mg$_\mathregular{Pt}$/cm$^2$"
    "\nIC 0.8, N212"
)

def build_grid(sample_list, group_name):
    ncols = len(CONDITIONS)
    nrows = len(sample_list)
    fs    = 7

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(2.75 * ncols, 2.4 * nrows),
        dpi=300,
        constrained_layout=True,
    )

    for row_idx, samp in enumerate(sample_list):
        for col_idx, (cond_folder, dur_folder, cond_label) in enumerate(CONDITIONS):
            ax = axes[row_idx, col_idx]

            ymax = 1.18 if " BP" not in cond_folder else 1.2

            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None

            if fp is None:
                ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)
                ax.set_xlim(0, 3500); ax.set_ylim(0.2, ymax)
                continue

            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                erev = calc_erev(T_K, P_H2, P_O2)
                draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                           fs=fs, lw=0.9,
                           show_legend=(row_idx == 0 and col_idx == 0),
                           show_annot=(row_idx == 0 and col_idx == 0),
                           annot_str=ANNOT, ymax=ymax)
            except Exception as e:
                print(f"  ERROR {samp} / {cond_folder}: {e}")
                ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)

            # Column header (top row only)
            if row_idx == 0:
                ax.set_title(cond_label, fontsize=fs + 2, fontweight="bold", pad=4)

            # Row label (left column only)
            if col_idx == 0:
                ax.set_ylabel("Cell Voltage (V)", fontsize=fs, labelpad=4)
                ax.text(-0.22, 0.5, ROW_LABELS[samp],
                        transform=ax.transAxes,
                        fontsize=fs + 2, fontweight="bold",
                        ha="center", va="center",
                        rotation=90, clip_on=False)
            else:
                ax.set_ylabel("Cell Voltage (V)", fontsize=fs)

    out_dir = os.path.join(BASE, "Overpotential Graphs")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, f"Overpot_grid_{group_name}.png")
    fig.savefig(out, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {out}")

# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
build_grid(MAIN_SAMPLES,  "Main")
build_grid(OTHER_SAMPLES, "Other")
print("Done.")
