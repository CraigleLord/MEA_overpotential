"""Render reference-style 2-panel LSV+Power figures for the SI.

Inputs : LSV_comparison.xlsx (8 data sheets pre-built from 32 source xlsx)
Output : LSV_reference_style_png/<cond>_combined.png  (and individual panels)

Series (matches the reference figure):
    CN BM     - solid, teal-green
    KB BM     - solid, powder blue
    VC Polyol - dotted, black
    AB Polyol - dotted, amber

Conditions (one figure each):
    Air 0 bar BP, Air 1.5 bar BP, O2 0 bar BP, O2 1.5 bar BP
"""

from __future__ import annotations

from pathlib import Path
import sys
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import openpyxl

ROOT = Path(__file__).parent
SRC  = ROOT / "LSV_comparison.xlsx"
OUT_DIR = ROOT / "LSV_reference_style_png"
OUT_DIR.mkdir(exist_ok=True)

# (catalyst, sample, color RGB 0-1, linestyle, legend label)
SERIES = [
    ("CN", "BM",     (45/255, 160/255, 130/255), "-",  "CN BM"),
    ("KB", "BM",     (120/255, 175/255, 220/255),"-",  "KB BM"),
    ("VC", "Polyol", (0.0, 0.0, 0.0),            ":",  "VC Polyol"),
    ("AB", "Polyol", (245/255, 180/255, 30/255), ":",  "AB Polyol"),
]

# (condition key, gas, bar string, data block start col 1-indexed in data sheet)
CONDS = [
    ("Air_0bp",  "Air", "0",   1),
    ("Air_15bp", "Air", "1.5", 4),
    ("O2_0bp",   "O2",  "0",   7),
    ("O2_15bp",  "O2",  "1.5", 10),
]


def read_series(wb: openpyxl.Workbook, catalyst: str, sample: str, col: int):
    """Return (i, V) numpy arrays from a (catalyst,sample) data sheet, given the 1-indexed start column."""
    sheet_name = f"{catalyst}_{sample}"
    ws = wb[sheet_name]
    i_vals, v_vals = [], []
    # Data starts at row 3.
    for r in range(3, ws.max_row + 1):
        a = ws.cell(row=r, column=col).value
        b = ws.cell(row=r, column=col + 1).value
        if a is None or b is None:
            continue
        try:
            ai = float(a); bi = float(b)
        except (TypeError, ValueError):
            continue
        # Filter the COM sentinel value if any leaked into the cells.
        if ai < -1e9 or bi < -1e9:
            continue
        i_vals.append(ai)
        v_vals.append(bi)
    return np.asarray(i_vals), np.asarray(v_vals)


def make_figure(data_for_series: dict, gas: str, bar: str, out_path_combined: Path):
    """Create one 2-panel figure (LSV + Power) for a single condition."""
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.weight": "bold",
        "axes.labelweight": "bold",
        "axes.linewidth": 1.4,
        "xtick.direction": "in",
        "ytick.direction": "in",
        "xtick.major.width": 1.2,
        "ytick.major.width": 1.2,
        "xtick.minor.width": 0.8,
        "ytick.minor.width": 0.8,
        "xtick.major.size": 5,
        "ytick.major.size": 5,
        "xtick.minor.size": 3,
        "ytick.minor.size": 3,
    })

    fig, (axL, axR) = plt.subplots(1, 2, figsize=(11.5, 4.4))
    fig.subplots_adjust(left=0.075, right=0.985, bottom=0.155, top=0.965, wspace=0.25)

    # Plot lines on both panels.
    for cat, sample, color, ls, label in SERIES:
        I, V = data_for_series[(cat, sample)]
        if I.size == 0:
            continue
        P = I * V  # mW/cm^2

        # LSV (left): V vs i
        line_kw = dict(color=color, linestyle=ls, linewidth=2.2, solid_capstyle="round",
                       dash_capstyle="round")
        if ls == ":":
            # Make dotted lines clearly round-dotted.
            line_kw["dashes"] = (1, 1.6)
            line_kw["linestyle"] = (0, line_kw["dashes"])
            del line_kw["dashes"]
        axL.plot(I, V, label=label, **line_kw)

        # Mark V at peak power on the LSV panel.
        idx = int(np.argmax(P))
        axL.plot([I[idx]], [V[idx]],
                 marker="o", markersize=8, markerfacecolor=color,
                 markeredgecolor=color, markeredgewidth=1.0,
                 linestyle="None", zorder=5)

        # Power (right): P vs i
        axR.plot(I, P, **line_kw)

    # Axis ranges and ticks (fixed per reference image).
    for ax in (axL, axR):
        ax.set_xlim(0, 3500)
        ax.set_xticks(np.arange(0, 3501, 500))
        ax.tick_params(axis="both", which="major", labelsize=11)
        for spine in ax.spines.values():
            spine.set_linewidth(1.4)
        ax.minorticks_off()
        ax.grid(False)

    axL.set_ylim(0.2, 1.2)
    axL.set_yticks(np.arange(0.2, 1.21, 0.2))
    axR.set_ylim(0, 1200)
    axR.set_yticks(np.arange(0, 1201, 200))

    # Axis labels (bold).
    axL.set_xlabel("Current Density (mAcm$^{-2}$)", fontsize=13, fontweight="bold")
    axL.set_ylabel("Cell Voltage (V)",              fontsize=13, fontweight="bold")
    axR.set_xlabel("Current Density (mAcm$^{-2}$)", fontsize=13, fontweight="bold")
    axR.set_ylabel("Power Density (mWcm$^{-2}$)",   fontsize=13, fontweight="bold")

    # Legend (top-left of LSV panel) - lines only, no markers.
    legend_handles = [
        Line2D([0], [0], color=c, linestyle=ls, linewidth=2.2, label=lab)
        for _, _, c, ls, lab in SERIES
    ]
    # Replace dotted legend handles with the same round-dot style.
    for h, (_, _, _, ls, _) in zip(legend_handles, SERIES):
        if ls == ":":
            h.set_linestyle((0, (1, 1.6)))
    axL.legend(handles=legend_handles, loc="upper left", frameon=False,
               fontsize=10.5, handlelength=2.4, labelspacing=0.35,
               bbox_to_anchor=(0.005, 0.99))

    # Condition annotation box (top-right of LSV panel).
    gas_label = "H$_2$/Air" if gas == "Air" else "H$_2$/O$_2$"
    annot_lines = [
        gas_label,
        "RH100",
        "Pt 5 wt%",
        r"0.05 mg$_{\mathrm{Pt}}$/cm$^{2}$",
        "IC 0.8, N212",
        f"{bar} bar$_{{g}}$",
    ]
    axL.text(0.985, 0.985, "\n".join(annot_lines),
             transform=axL.transAxes, ha="right", va="top",
             fontsize=10.5, fontweight="bold", linespacing=1.35)

    fig.savefig(out_path_combined, dpi=300, bbox_inches="tight", facecolor="white")

    # Save individual panels too.
    out_lsv = out_path_combined.with_name(out_path_combined.stem + "_LSV.png")
    out_pow = out_path_combined.with_name(out_path_combined.stem + "_Power.png")
    extentL = axL.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
    extentR = axR.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
    fig.savefig(out_lsv, dpi=300, bbox_inches=extentL.expanded(1.04, 1.04), facecolor="white")
    fig.savefig(out_pow, dpi=300, bbox_inches=extentR.expanded(1.04, 1.04), facecolor="white")
    plt.close(fig)


def main():
    if not SRC.exists():
        sys.exit(f"Input not found: {SRC}")
    print(f"Reading: {SRC}")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # Pre-load every needed series once (4 series x 4 conditions = 16 reads).
    data = {}  # (cat, sample, cond_key) -> (I, V)
    print("Loading series...")
    for cat, sample, _, _, _ in SERIES:
        for cond_key, _, _, col in CONDS:
            I, V = read_series(wb, cat, sample, col)
            data[(cat, sample, cond_key)] = (I, V)
            P = I * V
            if P.size:
                idx = int(np.argmax(P))
                print(f"  {cat:2s} {sample:6s} {cond_key:9s}: n={P.size:4d}, "
                      f"P_max={P[idx]:7.1f} mW/cm^2 at i={I[idx]:7.1f}, V={V[idx]:.3f}")

    # Render one figure per condition.
    for cond_key, gas, bar, _ in CONDS:
        d = {(cat, sample): data[(cat, sample, cond_key)] for cat, sample, *_ in SERIES}
        out = OUT_DIR / f"{cond_key}_combined.png"
        make_figure(d, gas, bar, out)
        print(f"  -> {out.name}")

    print(f"\nDone. Output dir: {OUT_DIR}")


if __name__ == "__main__":
    main()
