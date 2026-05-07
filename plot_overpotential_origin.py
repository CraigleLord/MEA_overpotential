"""
Overpotential breakdown → Origin .opju project file.

Creates fill-between graphs (Kinetic / Ohmic / Proton / Mass Transport)
that are fully editable in OriginLab 2019+.

Requirements:
    pip install originpro
    OriginLab 2019 or later must be installed.

Output:
    Durability I-V figure/tafel plot alt/overpotential decoupled/Overpot_breakdown.opju
"""

import os, sys, math
import numpy as np
import openpyxl
import originpro as op

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE     = os.path.dirname(os.path.abspath(__file__))
ALT_DIR  = os.path.join(BASE, "Durability I-V figure", "tafel plot alt")
ALT_DATA = os.path.join(ALT_DIR, "Overpotential DATA raw alt")
OUT_DIR  = os.path.join(ALT_DIR, "overpotential decoupled")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_OPJ  = os.path.join(OUT_DIR, "Overpot_breakdown.opju")

# ---------------------------------------------------------------------------
# Sample / condition definitions  (identical to plot_overpotential_all_fixed.py)
# ---------------------------------------------------------------------------
MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",     "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol", "AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol", "KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",    "AB BM":     "AB-BM",
}

DURS = [("BOL", "BOL"), ("30K", "30K"), ("75K", "75K")]

ALL_CONDITIONS = [
    ("Air",    "Air_0bp",  "Air 0 BP",    False),
    ("Air BP", "Air_15bp", "Air 1.5 BP",  True),
    ("O2",     "O2_0bp",   "O2 0 BP",     False),
    ("O2 BP",  "O2_15bp",  "O2 1.5 BP",   True),
]

BOL_CONDITIONS = [
    ("Air",    "BOL", "Air 0 BP"),
    ("Air BP", "BOL", "Air 1.5 BP"),
    ("O2",     "BOL", "O2 0 BP"),
    ("O2 BP",  "BOL", "O2 1.5 BP"),
]

FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

# ---------------------------------------------------------------------------
# Fill colors  (R, G, B)  — same hues as the matplotlib version
# ---------------------------------------------------------------------------
C_KIN  = (255, 160, 122)   # #FFA07A  salmon
C_OHM  = (144, 192, 144)   # #90C090  green
C_PRO  = (144, 144, 216)   # #9090D8  lavender
C_MASS = (240, 232,  96)   # #F0E860  yellow

# Origin encodes colors as R + G*256 + B*65536
def _oc(r, g, b):
    return r + g * 256 + b * 65536

OC_KIN  = _oc(*C_KIN)
OC_OHM  = _oc(*C_OHM)
OC_PRO  = _oc(*C_PRO)
OC_MASS = _oc(*C_MASS)
OC_BLK  = _oc(0, 0, 0)

# ---------------------------------------------------------------------------
# Data I/O helpers  (identical to original)
# ---------------------------------------------------------------------------
def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))
    if len(rows) < 5:
        raise ValueError("Too few data rows")
    wb.close()
    I_mA  = np.array([r[0] * 200 for r in rows])
    V     = np.array([r[1]        for r in rows])
    R_ohm = np.array([r[2]        for r in rows])
    R_pro = np.array([r[3]        for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None

# ---------------------------------------------------------------------------
# Boundary computation  (identical math to original)
# ---------------------------------------------------------------------------
def compute_boundaries(I_mA, V, R_ohm, R_pro, erev, intercept, slope):
    """Return (I, V_meas, V_erev, V_kin, V_ohm, V_pro) clipped to measured I-V."""
    idx_peak = int(np.argmax(I_mA))
    I  = I_mA[:idx_peak + 1]
    Vm = V[:idx_peak + 1]
    Ro = R_ohm[:idx_peak + 1]
    Rp = R_pro[:idx_peak + 1]

    I_A  = I / 200
    ln_j = np.log(I / 1000)
    eta_kin = -(intercept + slope * ln_j)

    V_erev = np.full_like(I, erev)
    V_kin  = np.maximum(erev - eta_kin, Vm)
    V_ohm  = np.maximum(V_kin - Ro * I_A, Vm)
    V_pro  = np.maximum(V_ohm - Rp * I_A, Vm)

    return I, Vm, V_erev, V_kin, V_ohm, V_pro

# ---------------------------------------------------------------------------
# Origin worksheet helper
# ---------------------------------------------------------------------------
# Column layout in every sheet:
#   col(A) [X]  = I_mA_cm2       current density
#   col(B) [Y]  = V_meas         measured cell voltage
#   col(C) [Y]  = V_erev         equilibrium voltage (Nernst)
#   col(D) [Y]  = V_kin          kinetic lower boundary
#   col(E) [Y]  = V_ohm          ohmic lower boundary
#   col(F) [Y]  = V_pro          proton lower boundary
#
# Fill regions for Origin "Fill Area Between Curves":
#   Kinetic       : C (upper) → D (lower)
#   Ohmic         : D (upper) → E (lower)
#   Proton        : E (upper) → F (lower)
#   Mass Transport: F (upper) → B (lower)

COL_NAMES = ['I_mA_cm2', 'V_meas', 'V_erev', 'V_kin', 'V_ohm', 'V_pro']
COL_UNITS = ['mA cm-2',  'V',      'V',      'V',     'V',     'V'     ]

def ensure_sheet(book, idx):
    """Return sheet at index idx, creating blank sheets as needed."""
    while len(book) <= idx:
        book.add_sheet()
    return book[idx]


def fill_sheet(wks, name, I, Vm, V_erev, V_kin, V_ohm, V_pro):
    """Write boundary data into an Origin worksheet."""
    wks.name = name[:31]
    data_arrays = [I, Vm, V_erev, V_kin, V_ohm, V_pro]
    wks.cols = len(data_arrays)
    for ci, (arr, lname, unit) in enumerate(zip(data_arrays, COL_NAMES, COL_UNITS)):
        col = wks[ci]
        col.lname = lname
        col.units = unit
        col.type  = 'X' if ci == 0 else 'Y'
        wks.from_list(ci, arr.tolist())

# ---------------------------------------------------------------------------
# Origin graph helper  — creates one graph page for a multi-panel layout
# ---------------------------------------------------------------------------
# Each panel is one Layer in Origin. We set:
#   - fill-area plots for the 4 overpotential regions
#   - a black line for the measured I-V curve
# LabTalk is used for fill-area-between-curves (not directly exposed via
# the originpro Python API).

_AXIS_SETTINGS = """
layer.x.from = 0;
layer.x.to   = 3500;
layer.y.from = 0.2;
layer.y.to   = 1.2;
layer.x.inc  = 1000;
layer.y.inc  = 0.2;
layer -s;
"""

def _add_fill_region(book_name, sheet_name, col_upper, col_lower, origin_color):
    """
    LabTalk snippet that adds one fill-area-between-curves region to the
    active Origin layer.

    col_upper / col_lower are 1-based column indices (A=1, B=2, …).
    origin_color is the Origin integer-encoded color.
    """
    # plotxy with a 3-column range: (X, Y_upper, Y_lower) → Fill Area Between
    return (
        f"plotxy ([{book_name}]{sheet_name}!"
        f"(col(A),col({col_upper}),col({col_lower}))) "
        f"type:=203 color:={origin_color};\n"
        # type 203 = "Fill Area Between Two Datasets" in Origin 2019+
    )


def build_layer(book_name, sheet_name, row, col, nrows, ncols,
                row_label="", col_label="", is_topleft=False):
    """
    Add a layer to the current graph page at grid position (row, col),
    populate it with fill plots, and return the layer index.

    Coordinate system: layers are arranged in a nrows × ncols grid.
    Origin layer positions are set by percentage (0-100) of page size.
    """
    # Page margins (percent of page)
    left_margin  = 12.0 if ncols > 1 else 8.0
    top_margin   = 8.0  if nrows > 1 else 8.0
    right_margin = 3.0
    bot_margin   = 8.0

    usable_w = 100.0 - left_margin - right_margin
    usable_h = 100.0 - top_margin  - bot_margin
    gap_x    = 4.0
    gap_y    = 5.0

    cell_w = (usable_w - gap_x * (ncols - 1)) / ncols
    cell_h = (usable_h - gap_y * (nrows - 1)) / nrows

    x0 = left_margin + col * (cell_w + gap_x)
    y0 = top_margin  + row * (cell_h + gap_y)

    lt_add = (
        f"layer -a;\n"            # add new layer
        f"layer.x.left  = {x0:.2f};\n"
        f"layer.y.top   = {y0:.2f};\n"
        f"layer.x.right = {x0 + cell_w:.2f};\n"
        f"layer.y.bottom = {y0 + cell_h:.2f};\n"
    )
    op.run(lt_add)

    # Fill the 4 overpotential regions
    fills = [
        (3, 4, OC_KIN,  "Kinetic"),
        (4, 5, OC_OHM,  "Ohmic"),
        (5, 6, OC_PRO,  "Proton"),
        (6, 2, OC_MASS, "Mass Transport"),
    ]
    for col_up, col_lo, oc, _ in fills:
        op.run(_add_fill_region(book_name, sheet_name, col_up, col_lo, oc))

    # Black I-V line on top
    op.run(
        f"plotxy ([{book_name}]{sheet_name}!(col(A),col(B))) "
        f"type:=200 color:={OC_BLK};\n"
    )

    # Axis ranges
    op.run(_AXIS_SETTINGS)

    # Suppress tick labels except on outermost panels
    show_x_label = (row == nrows - 1)
    show_y_label = (col == 0)
    if not show_x_label:
        op.run("xaxis -t 0;\n")    # hide x tick labels
    if not show_y_label:
        op.run("yaxis -t 0;\n")    # hide y tick labels

    # Row / column labels as title text objects
    if col_label and row == 0:
        op.run(
            f'label -n col_lbl_{col} -xa {x0 + cell_w/2:.2f} '
            f'-ya {y0 - 3:.2f} -s -h "{col_label}";\n'
        )
    if row_label and col == 0:
        op.run(
            f'label -n row_lbl_{row} -xa {x0 - 5:.2f} '
            f'-ya {y0 + cell_h/2:.2f} -s -r 90 -h "{row_label}";\n'
        )


def new_graph_page(page_name, nrows, ncols):
    """Create a blank graph page and return it."""
    graph = op.new_graph(lname=page_name)
    # Remove the default layer that Origin always adds
    op.run("layer -d 1;\n")
    return graph

# ---------------------------------------------------------------------------
# Figure builders
# ---------------------------------------------------------------------------

def build_bol_grid(sample_list, group_name):
    """BOL-only 4 × 4 grid: rows = samples, cols = conditions."""
    nrows = len(sample_list)
    ncols = len(BOL_CONDITIONS)
    page_name = f"BOL_grid_{group_name}"
    print(f"\nBuilding {page_name} …")

    # ── workbook ──────────────────────────────────────────────────────────
    book = op.new_book('w', lname=f"Data_BOL_{group_name}")
    sheet_idx = 0

    # collect (wks, row, col, col_label, row_label) for panels with data
    panels = []

    for ri, samp in enumerate(sample_list):
        for ci, (cond_folder, dur_folder, cond_label) in enumerate(BOL_CONDITIONS):
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None
            if fp is None:
                panels.append(None)
                continue
            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                erev = calc_erev(T_K, P_H2, P_O2)
                I, Vm, V_erev, V_kin, V_ohm, V_pro = compute_boundaries(
                    I_mA, V, R_ohm, R_pro, erev, intercept, slope)
                wks = ensure_sheet(book, sheet_idx)
                fill_sheet(wks, f"{LABELS[samp]}_{cond_folder}", I, Vm, V_erev, V_kin, V_ohm, V_pro)
                panels.append((book.name, wks.name, ri, ci,
                               cond_label, LABELS[samp]))
                sheet_idx += 1
                print(f"  OK  {LABELS[samp]} / {cond_folder}")
            except Exception as exc:
                print(f"  ERR {LABELS[samp]} / {cond_folder}: {exc}")
                panels.append(None)

    # ── graph page ─────────────────────────────────────────────────────────
    graph = op.new_graph(lname=page_name)
    op.run("layer -d 1;\n")   # remove default layer

    for item in panels:
        if item is None:
            continue
        book_name, sheet_name, ri, ci, col_lbl, row_lbl = item
        build_layer(book_name, sheet_name,
                    ri, ci, nrows, ncols,
                    row_label=row_lbl, col_label=col_lbl)

    print(f"  Graph page '{page_name}' created.")


def build_dur_grid(sample_list, cond_folder, safe_name, group_name):
    """Durability grid: rows = BOL/30K/75K, cols = samples."""
    nrows = len(DURS)
    ncols = len(sample_list)
    page_name = f"Dur_{safe_name}_{group_name}"
    print(f"\nBuilding {page_name} …")

    book = op.new_book('w', lname=f"Data_{safe_name}_{group_name}")
    sheet_idx = 0
    panels = []

    for ri, (dur_label, dur_folder) in enumerate(DURS):
        for ci, samp in enumerate(sample_list):
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None
            if fp is None:
                panels.append(None)
                continue
            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                erev = calc_erev(T_K, P_H2, P_O2)
                I, Vm, V_erev, V_kin, V_ohm, V_pro = compute_boundaries(
                    I_mA, V, R_ohm, R_pro, erev, intercept, slope)
                wks = ensure_sheet(book, sheet_idx)
                fill_sheet(wks, f"{dur_label}_{LABELS[samp]}", I, Vm, V_erev, V_kin, V_ohm, V_pro)
                panels.append((book.name, wks.name, ri, ci,
                               LABELS[samp], dur_label))
                sheet_idx += 1
                print(f"  OK  {dur_label} / {LABELS[samp]}")
            except Exception as exc:
                print(f"  ERR {dur_label} / {LABELS[samp]}: {exc}")
                panels.append(None)

    graph = op.new_graph(lname=page_name)
    op.run("layer -d 1;\n")

    for item in panels:
        if item is None:
            continue
        book_name, sheet_name, ri, ci, col_lbl, row_lbl = item
        build_layer(book_name, sheet_name,
                    ri, ci, nrows, ncols,
                    row_label=row_lbl, col_label=col_lbl)

    print(f"  Graph page '{page_name}' created.")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # Connect to (or launch) Origin
    op.set_show(True)          # show Origin window while building
    op.new()                   # fresh project

    # BOL grids
    build_bol_grid(MAIN_SAMPLES,  "Main")
    build_bol_grid(OTHER_SAMPLES, "Other")

    # Durability grids
    for cond_folder, safe_name, _, _ in ALL_CONDITIONS:
        build_dur_grid(MAIN_SAMPLES,  cond_folder, safe_name, "Main")
        build_dur_grid(OTHER_SAMPLES, cond_folder, safe_name, "Other")

    # Save
    op.save(OUT_OPJ)
    print(f"\nSaved: {OUT_OPJ}")
    op.set_show(False)
