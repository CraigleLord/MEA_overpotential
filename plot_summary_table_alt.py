"""
Summary tables using alt Tafel kinetics (window -3.5 to -2.0).
Reads slope/intercept from: Durability I-V figure/tafel plot alt/Kinetic slope and intercept alt.xlsx
Reads HFR/Proton from:      HFR+Proton All.xlsx  (unchanged)
Output: Durability I-V figure/tafel plot alt/Table_O2_15bp_Main_alt.png
        Durability I-V figure/tafel plot alt/Table_O2_15bp_Other_alt.png
"""

import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
ALT_DIR = os.path.join(BASE, "Durability I-V figure", "tafel plot alt")

MAIN_SAMPLES  = ["CN BM", "KB BM", "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM", "AB BM"]
DURS          = ["BOL", "30K", "75K"]

DISPLAY_NAMES = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

# ---------------------------------------------------------------------------
# Data reading
# ---------------------------------------------------------------------------

def read_kinetics():
    path = os.path.join(ALT_DIR, "Kinetic slope and intercept alt NEW.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    result = {}
    current_samp = None
    for row in rows:
        samp, dur, slope, intercept = row[0], row[1], row[2], row[3]
        if samp is not None:
            current_samp = samp
        if dur in ("BOL", "30K", "75K") and current_samp is not None:
            result[(current_samp, dur)] = (
                float(slope)     if isinstance(slope,     (int, float)) else np.nan,
                float(intercept) if isinstance(intercept, (int, float)) else np.nan,
            )
    return result


def read_hfr_proton(sheet_name, sample_list):
    wb = openpyxl.load_workbook(
        os.path.join(BASE, "HFR+Proton All.xlsx"), data_only=True
    )
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    hfr_row = pro_row = None
    for i, row in enumerate(rows):
        if row[0] == "HFR 1.5 bar BP":
            hfr_row = i
        elif row[0] == "Proton Transfer Resistance":
            pro_row = i
    wb.close()

    result = {}
    dur_col_offset = {"BOL": 1, "30K": 3, "75K": 5}
    for samp_idx, samp in enumerate(sample_list):
        base_col = samp_idx * 6
        for dur in DURS:
            col = base_col + dur_col_offset[dur]
            hfr = rows[hfr_row][col] if hfr_row is not None else np.nan
            pro = rows[pro_row][col] if pro_row is not None else np.nan
            result[(samp, dur)] = (
                float(hfr) if isinstance(hfr, (int, float)) else np.nan,
                float(pro) if isinstance(pro, (int, float)) else np.nan,
            )
    return result

# ---------------------------------------------------------------------------
# Table styling
# ---------------------------------------------------------------------------

HDR_BG  = "#1E3A5F"
HDR_FG  = "white"
ROW_A   = "#D9E4F0"
ROW_B   = "#FFFFFF"
SAMP_FG = "#1E3A5F"

COL_LABELS = ["Sample", "Dur.", "Tafel Slope", "Tafel Intercept"]
COL_W      = [0.14, 0.07, 0.22, 0.22]


def fmt(v, decimals=3):
    return f"{v:.{decimals}f}" if (v is not None and not np.isnan(v)) else "—"


def build_table(sample_list, group_name, sheet_name):
    kinetics = read_kinetics()

    rows_data = []
    for samp in sample_list:
        for i, dur in enumerate(DURS):
            b, a = kinetics.get((samp, dur), (np.nan, np.nan))
            samp_label = DISPLAY_NAMES[samp] if i == 0 else ""
            rows_data.append((samp_label, dur, b, a))

    cell_text = []
    for samp_label, dur, b, a in rows_data:
        cell_text.append([
            samp_label,
            dur,
            fmt(b, 4),
            fmt(a, 3),
        ])

    n_data_rows = len(cell_text)
    n_cols      = len(COL_LABELS)

    data_row_h_in = 0.34
    hdr_row_h_in  = 0.44
    fig_h = n_data_rows * data_row_h_in + hdr_row_h_in + 0.70
    fig_w = 7.0

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    tbl = ax.table(
        cellText=cell_text,
        colLabels=COL_LABELS,
        loc="upper center",
        cellLoc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9.5)

    hdr_h_frac  = hdr_row_h_in  / fig_h
    data_h_frac = data_row_h_in / fig_h

    for j in range(n_cols):
        tbl[0, j].set_height(hdr_h_frac)
    for i in range(1, n_data_rows + 1):
        for j in range(n_cols):
            tbl[i, j].set_height(data_h_frac)

    total_w = sum(COL_W)
    for j, w in enumerate(COL_W):
        for i in range(n_data_rows + 1):
            tbl[i, j].set_width(w / total_w)

    # Header styling
    for j in range(n_cols):
        cell = tbl[0, j]
        cell.set_facecolor(HDR_BG)
        cell.set_text_props(color=HDR_FG, fontweight="bold")
        cell.set_edgecolor("white")

    # Data row styling
    for i, (samp_label, dur, *_) in enumerate(rows_data):
        tbl_row   = i + 1
        group_idx = i // 3
        bg = ROW_A if group_idx % 2 == 0 else ROW_B

        for j in range(n_cols):
            cell = tbl[tbl_row, j]
            cell.set_facecolor(bg)
            cell.set_edgecolor("#AAAAAA")

        if samp_label:
            tbl[tbl_row, 0].set_text_props(color=SAMP_FG, fontweight="bold")

        if i % 3 == 0 and i > 0:
            for j in range(n_cols):
                tbl[tbl_row, j].set_edgecolor("#555555")

    fig.text(0.5, 0.98,
             f"O₂ 1.5 bar BP  —  {group_name} samples  (alt window −3.0 to −2.0)",
             ha="center", va="top",
             fontsize=12, fontweight="bold", color=HDR_BG)

    out = os.path.join(BASE, "260603 New data set (reproducibility)", "Final Plots",
                       f"Table_O2_15bp_{group_name}_alt.png")
    fig.savefig(out, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out}")


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
build_table(MAIN_SAMPLES,  "Main",  "Main")
build_table(OTHER_SAMPLES, "Other", "Other")
print("Done.")
