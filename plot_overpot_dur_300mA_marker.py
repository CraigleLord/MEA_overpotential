"""
All Overpot_dur_* figures with red markers at 300 mA/cm²
(instead of the 500 mA/cm² used in the originals).
Format identical to plot_overpotential_all_fixed.py.
Output: Figure Plots (Claude AI)/Overpot_dur_*_300mA.png
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE     = os.path.dirname(os.path.abspath(__file__))
ALT_DATA = os.path.join(BASE, "Durability I-V figure", "tafel plot alt",
                         "Overpotential DATA raw alt")
OUT_DIR        = os.path.join(BASE, "Figure Plots (Claude AI)")
OUT_DIR_260603 = os.path.join(BASE, "260603 New data set (reproducibility)")
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(OUT_DIR_260603, exist_ok=True)

MAIN_SAMPLES  = ["CN BM",     "KB BM",     "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM",     "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

DURS = [("BOL", "BOL"), ("75K", "75K")]

CONDITIONS = [
    ("Air",    "Air_0bp",  "Air 0 BP",       False),
    ("Air BP", "Air_15bp", "Air 1.5 BP",     True),
    ("O2",     "O2_0bp",   r"O$_2$ 0 BP",    False),
    ("O2 BP",  "O2_15bp",  r"O$_2$ 1.5 BP",  True),
]

FILE_STEMS = {
    ("Air",    "CN BM"):     "CN BM air",
    ("Air",    "KB BM"):     "KB BM air",
    ("Air",    "VC Polyol"): "VC Polyol air",
    ("Air",    "AB Polyol"): "AB Polyol air",
    ("Air",    "CN Polyol"): "CN Polyol air",
    ("Air",    "KB Polyol"): "KB Polyol air",
    ("Air",    "VC BM"):     "VC BM air",
    ("Air",    "AB BM"):     "AB BM Air",

    ("Air BP", "CN BM"):     "CN BM 1.5bp air",
    ("Air BP", "KB BM"):     "KB BM 1.5bp air",
    ("Air BP", "VC Polyol"): "VC Polyol 1.5bp air",
    ("Air BP", "AB Polyol"): "AB Polyol 1.5bp air",
    ("Air BP", "CN Polyol"): "CN Polyol 1.5bp air",
    ("Air BP", "KB Polyol"): "KB Polyol 1.5bp air",
    ("Air BP", "VC BM"):     "VC BM 1.5bp air",
    ("Air BP", "AB BM"):     "AB BM 1.5bp air",

    ("O2",    "CN BM"):      "CN BM o2",
    ("O2",    "KB BM"):      "KB BM O2",
    ("O2",    "VC Polyol"):  "VC Polyol O2",
    ("O2",    "AB Polyol"):  "AB Polyol O2",
    ("O2",    "CN Polyol"):  "CN Polyol O2",
    ("O2",    "KB Polyol"):  "KB Polyol O2",
    ("O2",    "VC BM"):      "VC BM O2",
    ("O2",    "AB BM"):      "AB BM  O2",

    ("O2 BP", "CN BM"):      "CN BM 1.5bp O2",
    ("O2 BP", "KB BM"):      "KB BM 1.5bp O2",
    ("O2 BP", "VC Polyol"):  "VC Polyol 1.5bp O2",
    ("O2 BP", "AB Polyol"):  "AB Polyol 1.5bp O2",
    ("O2 BP", "CN Polyol"):  "CN Polyol 1.5bp O2",
    ("O2 BP", "KB Polyol"):  "KB Polyol 1.5bp O2",
    ("O2 BP", "VC BM"):      "VC BM 1.5bp O2",
    ("O2 BP", "AB BM"):      "AB BM 1.5bp O2",
}

C_KIN  = "#FFA07A"
C_OHM  = "#90C090"
C_PRO  = "#9090D8"
C_MASS = "#F0E860"

MARKER_I = 500.0   # mA/cm²

# EIS-derived proton resistance overrides (replaces per-point Excel column D)
EIS_RPRO = {
    "CN BM": 0.020229,
}

# 260603 reproducibility data overrides: (cond_folder, stem) -> folder containing
# {dur_folder}/{stem}_edited.xlsx (dur_folder = "BOL"/"30K"/"75K")
REPRO_DATA_260603 = os.path.join(BASE, "260603 New data set (reproducibility)",
                                  "Overpotential DATA raw alt")
DATA_OVERRIDES = {
    ("Air BP", "KB BM 1.5bp air"): os.path.join(REPRO_DATA_260603, "Air BP", "KB BM"),
}


def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]
    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value
    for v in (T_K, P_H2, P_O2, intercept, slope):
        if not isinstance(v, (int, float)):
            raise ValueError(f"Non-numeric parameter: {v!r}")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, _, d, e = row[0], row[1], row[2], row[3], row[4]
        if isinstance(a, (int, float)) and a > 0 and isinstance(b, (int, float)):
            rows.append((a, b,
                         float(d) if isinstance(d, (int, float)) else np.nan,
                         float(e) if isinstance(e, (int, float)) else np.nan))
    if len(rows) < 5:
        raise ValueError("Too few data rows")
    wb.close()
    I_mA  = np.array([r[0] * 200 for r in rows])
    V     = np.array([r[1]        for r in rows])
    R_ohm = np.array([r[2]        for r in rows])
    R_pro = np.array([r[3]        for r in rows])
    return I_mA, V, R_ohm, R_pro, float(T_K), float(P_H2), float(P_O2), float(intercept), float(slope)


def find_file(cond_folder, dur_folder, stem):
    override = DATA_OVERRIDES.get((cond_folder, stem))
    if override:
        for dur_try in (dur_folder, dur_folder.lower()):
            for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
                p = os.path.join(override, dur_try, stem + sfx)
                if os.path.exists(p):
                    return p
    for dur_try in (dur_folder, dur_folder.lower()):
        folder = os.path.join(ALT_DATA, cond_folder, dur_try, "Edited")
        for sfx in ("_edited.xlsx", "_unchanged.xlsx"):
            p = os.path.join(folder, stem + sfx)
            if os.path.exists(p):
                return p
    return None


def draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
               fs=18, lw=1.6, show_legend=False, ymax=1.18,
               legend_frameon=False, xlim=(0, 3.5), xtick_step=1):

    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    I_A  = I_mA / 200
    ln_j = np.log(I_mA / 1000)

    eta_kin = -(intercept + slope * ln_j)
    eta_ohm = R_ohm * I_A
    eta_pro = R_pro * I_A

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm_ = V_kin - eta_ohm
    V_pro_ = V_ohm_ - eta_pro

    V_kin_c  = np.maximum(V_kin,  V)
    V_ohm_c  = np.maximum(V_ohm_, V)
    V_pro_c  = np.maximum(V_pro_, V)

    I_plot = I_mA / 1000  # convert to A cm⁻² for display

    ax.fill_between(I_plot, V_kin_c,  V_erev,   color=C_KIN,  alpha=0.85, label="Kinetic")
    ax.fill_between(I_plot, V_ohm_c,  V_kin_c,  color=C_OHM,  alpha=0.85, label="Ohmic")
    ax.fill_between(I_plot, V_pro_c,  V_ohm_c,  color=C_PRO,  alpha=0.85, label="Proton")
    ax.fill_between(I_plot, V,        V_pro_c,  color=C_MASS, alpha=0.85, label="Mass Transport")

    ax.plot(I_plot, V, color="black", linewidth=lw, zorder=5)

    i_m      = MARKER_I
    v_meas_m = float(np.interp(i_m, I_mA, V))
    ax.plot(i_m / 1000, v_meas_m, marker="o", ms=6, color="red", linestyle="none", zorder=10)
    ax.text(i_m / 1000 + 0.1, v_meas_m, f"{v_meas_m:.2f}",
            color="red", fontsize=fs, va="center", ha="left")

    ax.set_xlim(*xlim)
    ax.set_ylim(0.2, ymax)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    ax.xaxis.set_major_locator(plt.MultipleLocator(xtick_step))
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}"))
    ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)

    if show_legend:
        patches = [
            mpatches.Patch(color=C_KIN,  label="Kinetic"),
            mpatches.Patch(color=C_OHM,  label="Ohmic"),
            mpatches.Patch(color=C_PRO,  label="Proton"),
            mpatches.Patch(color=C_MASS, label="Mass Transport"),
        ]
        ax.legend(handles=patches, loc="upper right", fontsize=12,
                  frameon=legend_frameon, facecolor="none", edgecolor="black",
                  handlelength=1.0, borderpad=0.3, labelspacing=0.15)


def current_at_voltage(I_mA, V, v_target):
    """Current density (mA cm⁻²) where the I-V curve crosses v_target. NaN if out of range."""
    idx_peak = int(np.argmax(I_mA))
    I_t, V_t = I_mA[:idx_peak + 1], V[:idx_peak + 1]
    order = np.argsort(V_t)
    V_s, I_s = V_t[order], I_t[order]
    if v_target < V_s[0] or v_target > V_s[-1]:
        return np.nan
    return float(np.interp(v_target, V_s, I_s))


def calc_bar_values(I_mA, V, R_ohm, R_pro, erev, intercept, slope, i=None):
    """Overpotential components (V) at `i` mA/cm² (default MARKER_I)."""
    idx_peak = int(np.argmax(I_mA))
    I_mA  = I_mA[:idx_peak + 1]
    V     = V[:idx_peak + 1]
    R_ohm = R_ohm[:idx_peak + 1]
    R_pro = R_pro[:idx_peak + 1]

    if i is None:
        i = MARKER_I
    I_A_i   = i / 200
    ln_j_i  = np.log(i / 1000)
    Ro_i    = float(np.interp(i, I_mA, R_ohm))
    Rp_i    = float(np.interp(i, I_mA, R_pro))
    V_i     = float(np.interp(i, I_mA, V))

    eta_kin = max(0.0, -(intercept + slope * ln_j_i))
    eta_ohm = max(0.0, Ro_i * I_A_i)
    eta_pro = max(0.0, Rp_i * I_A_i)
    eta_mt  = max(0.0, erev - eta_kin - eta_ohm - eta_pro - V_i)
    return eta_kin, eta_ohm, eta_pro, eta_mt


_BAR_LABEL_X_OFFSET = 0.055   # figure fraction: fixed distance left of bar-axis left edge to label centre
_BAR_WSPACE         = 0.45    # column gap must exceed _BAR_LABEL_X_OFFSET + label half-width
_COMP_NAMES         = ["Kinetic", "Ohmic", "Proton", "Mass Transport"]


def draw_component_bar_panel(ax, bar_data, comp_idx, sample_list, y_max,
                              fs=18, show_legend=False, show_xtick_labels=True):
    """
    Waterfall stacked bar (one bar per sample, per CLAUDE.md convention):
      • bottom: BOL value  (light gray #cccccc)
      • top:    max(75K − BOL, 0)  degradation increment  (dark gray #555555)
    Total bar height = max(BOL, 75K) ≈ 75K for overpotential.
    bar_data: {dur_label: {samp: (kin, ohm, pro, mt) or None}}
    comp_idx: 0=Kinetic, 1=Ohmic, 2=Proton, 3=Mass Transport
    """
    n = len(sample_list)
    x = np.arange(n, dtype=float)
    w = 0.55   # per CLAUDE.md

    bol_vals = np.array([
        bar_data["BOL"][s][comp_idx] if bar_data["BOL"].get(s) is not None else 0.0
        for s in sample_list
    ])
    k75_vals = np.array([
        bar_data["75K"][s][comp_idx] if bar_data["75K"].get(s) is not None else 0.0
        for s in sample_list
    ])
    incr = np.maximum(k75_vals - bol_vals, 0.0)  # degradation increment (always ≥ 0)

    ax.bar(x, bol_vals, width=w, facecolor="#cccccc",
           edgecolor="black", linewidth=0.5, label="BOL")
    ax.bar(x, incr, width=w, bottom=bol_vals, facecolor="#555555",
           edgecolor="black", linewidth=0.5, label="75K")

    tick_labels = [LABELS[s].replace("-", "\n") for s in sample_list]
    ax.set_xticks(x)
    if show_xtick_labels:
        ax.set_xticklabels(tick_labels, fontsize=fs)
    else:
        ax.tick_params(labelbottom=False)
    ax.set_xlim(-0.6, n - 0.4)
    ax.set_ylim(0, y_max)
    # Explicit step + format chosen by range to prevent duplicate tick labels
    # (MaxNLocator can produce 0.05-spaced ticks that round to the same 1-dp string)
    if y_max <= 0.08:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.02))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max <= 0.20:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.05))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max <= 0.50:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.10))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
    else:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.20))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)
    if show_legend:
        ax.legend(fontsize=fs, frameon=False,
                  handlelength=1.2, borderpad=0.2, labelspacing=0.3)


def draw_grouped_bar_panel(ax, bar_data, comp_idx, sample_list, y_max,
                           fs=18, show_legend=False, show_xtick_labels=True,
                           y_step=None, y_min=0.0):
    """
    Side-by-side grouped bars: BOL (left, light gray) and 75K (right, dark gray) per sample.
    """
    n = len(sample_list)
    x = np.arange(n, dtype=float)
    w = 0.35

    bol_vals = np.array([
        bar_data["BOL"][s][comp_idx] if bar_data["BOL"].get(s) is not None else 0.0
        for s in sample_list
    ])
    k75_vals = np.array([
        bar_data["75K"][s][comp_idx] if bar_data["75K"].get(s) is not None else 0.0
        for s in sample_list
    ])

    ax.bar(x - w / 2, bol_vals, width=w, facecolor="#cccccc",
           edgecolor="black", linewidth=0.5, label="BOL")
    ax.bar(x + w / 2, k75_vals, width=w, facecolor="#555555",
           edgecolor="black", linewidth=0.5, label="75K")

    tick_labels = [LABELS[s].replace("-", "\n") for s in sample_list]
    ax.set_xticks(x)
    if show_xtick_labels:
        ax.set_xticklabels(tick_labels, fontsize=fs)
    else:
        ax.tick_params(labelbottom=False)
    ax.set_xlim(-0.6, n - 0.4)
    ax.set_ylim(y_min, y_max)
    if y_step is not None:
        dec = len(f"{y_step:.10f}".rstrip("0").split(".")[1])
        ax.yaxis.set_major_locator(plt.MultipleLocator(y_step))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.{dec}f}"))
    elif y_max <= 0.006:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.001))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.3f}"))
    elif y_max <= 0.015:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.002))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.3f}"))
    elif y_max <= 0.030:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.005))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.3f}"))
    elif y_max <= 0.08:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.02))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max <= 0.20:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.05))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max <= 0.50:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.10))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
    else:
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.20))
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=fs, width=1.2, length=5)
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)
    if show_legend:
        ax.legend(fontsize=fs, frameon=False,
                  handlelength=1.2, borderpad=0.2, labelspacing=0.3)


def _nice_upper(v):
    """Ceiling to a nice round number with ~25% proportional headroom."""
    if v <= 0:
        return 0.0
    padded = v * 1.25
    mag = 10 ** math.floor(math.log10(padded))
    for mult in (1.0, 2.0, 5.0):
        step = mult * mag
        candidate = math.ceil(padded / step) * step
        if candidate / step <= 5:
            return candidate
    return math.ceil(padded * 1000) / 1000


def build_dur_grid(sample_list, cond_folder, safe_name, is_15bp, group_name,
                   samples_as_rows=False, use_delta=False):
    if samples_as_rows:
        nrows, ncols = len(sample_list), len(DURS)
    else:
        nrows, ncols = len(DURS), len(sample_list)

    fs        = 18
    ymax      = 1.18
    title_pad = fs * 0.67
    panel_xlim  = (0, 2.2) if safe_name == "Air_15bp" else (0, 3.5)
    panel_xtick = 0.5      if safe_name == "Air_15bp" else 1

    # ── Left-margin layout (all values are figure fractions, 0–1) ───────────
    # Left-to-right order:  row labels │ y-axis title │ tick numbers │ axes
    # Rule: ROW_LABEL_X  <  YLABEL_X  <  LEFT  (leave ~0.04 gap for tick numbers)
    # Row labels use ax.get_position() for y, so changing LEFT never moves them.
    #
    # For samples_as_rows figures, one bar column (col 3) is appended using
    # GridSpec.  Each row of col 3 shows one overpotential component (grouped
    # BOL/75K bars).  IV-column positions are preserved by scaling the
    # left-margin fractions by  (IV width) / (total figure width).
    if samples_as_rows:
        FIG_W_IV  = 5.0    # per IV-column width (inches)
        FIG_W_BAR = 7.0    # bar-column width — wider than IV columns for x-label spacing
        ncols_bar = 1      # one column: 4 component subplots, one per row
        total_w   = FIG_W_IV * ncols + FIG_W_BAR * ncols_bar
        _s        = (FIG_W_IV * ncols) / total_w   # scale factor to preserve absolute positions
        ROW_LABEL_X  = 0.04 * _s
        YLABEL_X     = 0.11 * _s
        LEFT         = 0.20 * _s
        width_ratios = [FIG_W_IV] * ncols + [FIG_W_BAR] * ncols_bar
    else:
        FIG_W_IV  = 4.0
        ncols_bar = 0
        total_w   = FIG_W_IV * ncols
        ROW_LABEL_X  = 0.01
        YLABEL_X     = 0.04
        LEFT         = 0.11
        width_ratios = None
    XLABEL_Y = 0.06
    # ────────────────────────────────────────────────────────────────────────

    fig = plt.figure(figsize=(total_w, 3.5 * nrows), dpi=300)
    gs  = fig.add_gridspec(nrows, ncols + ncols_bar,
                            width_ratios=width_ratios if width_ratios else [1] * ncols)
    axes = np.array([[fig.add_subplot(gs[r, c]) for c in range(ncols)]
                     for r in range(nrows)])
    if ncols_bar:
        # One subplot per row in the bar column (each shows one component)
        bar_axes = [fig.add_subplot(gs[r, ncols]) for r in range(nrows)]
        bar_data = {dl: {} for dl, _ in DURS}
    else:
        bar_axes = []
    fig.subplots_adjust(left=LEFT, hspace=0.25, wspace=_BAR_WSPACE if ncols_bar else 0.25)
    # Centre supxlabel/supylabel over IV columns only (bar column must be excluded)
    iv_x0 = axes[0, 0].get_position().x0
    iv_x1 = axes[0, ncols - 1].get_position().x1
    iv_xc = (iv_x0 + iv_x1) / 2
    fig.supxlabel(r"Current Density (Acm$^{-2}$)", fontsize=fs, y=XLABEL_Y, x=iv_xc)
    fig.supylabel("Cell Voltage (V)", fontsize=fs, x=YLABEL_X)

    if samples_as_rows:
        rows = list(enumerate(sample_list))
        cols = list(enumerate(DURS))
    else:
        rows = [(i, d) for i, d in enumerate(DURS)]
        cols = list(enumerate(sample_list))

    for row_idx, row_key in rows:
        for col_idx, col_key in cols:
            if samples_as_rows:
                samp              = row_key
                dur_label, dur_folder = col_key
            else:
                dur_label, dur_folder = row_key
                samp              = col_key

            ax         = axes[row_idx, col_idx]
            is_topleft = (row_idx == 0 and col_idx == 0)
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None

            if fp is None:
                ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)
                ax.set_xlim(*panel_xlim); ax.set_ylim(0.2, ymax)
                if ncols_bar:
                    bar_data[dur_label][samp] = None
            else:
                try:
                    I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                    if samp in EIS_RPRO:
                        R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                    erev = calc_erev(T_K, P_H2, P_O2)
                    draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                               fs=fs, lw=1.6,
                               show_legend=is_topleft,
                               ymax=ymax,
                               legend_frameon=False,
                               xlim=panel_xlim,
                               xtick_step=panel_xtick)
                    if ncols_bar:
                        bar_data[dur_label][samp] = calc_bar_values(
                            I_mA, V, R_ohm, R_pro, erev, intercept, slope)
                except Exception as e:
                    print(f"  ERROR {samp}/{dur_label}/{cond_folder}: {e}")
                    ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                            ha="center", va="center", fontsize=fs)
                    if ncols_bar:
                        bar_data[dur_label][samp] = None

            if row_idx == 0:
                col_title = dur_label if samples_as_rows else LABELS[samp]
                ax.set_title(col_title, fontsize=fs, fontweight="bold", pad=title_pad)
            if col_idx == 0:
                row_label = LABELS[samp] if samples_as_rows else dur_label
                pos   = axes[row_idx, 0].get_position()
                y_fig = (pos.y0 + pos.y1) / 2
                fig.text(ROW_LABEL_X, y_fig, row_label,
                         fontsize=fs, fontweight="bold",
                         ha="center", va="center", rotation=90, clip_on=False)

    # ── Per-component bar subplots in col 3 (samples_as_rows only) ──────────
    if ncols_bar:
        if use_delta:
            # Grouped BOL/75K bars: shared per-component y-scale across all samples
            comp_ymaxes = [0.0] * 4
            for dl, _ in DURS:
                for s in sample_list:
                    vals = bar_data[dl].get(s)
                    if vals is not None:
                        for ci, v in enumerate(vals):
                            comp_ymaxes[ci] = max(comp_ymaxes[ci], v)
            comp_ymaxes = [max(_nice_upper(m), 0.01) for m in comp_ymaxes]
            # Per-panel y-tick step overrides for Air_15bp grids
            Y_STEP_OVERRIDES = {
                ("Main",  3): 0.02,   # Mass Transport
                ("Other", 2): 0.01,   # Proton
            }
            for comp_idx, ax_bar in enumerate(bar_axes):
                draw_grouped_bar_panel(
                    ax_bar, bar_data, comp_idx, sample_list,
                    y_max=comp_ymaxes[comp_idx], fs=fs,
                    show_legend=(comp_idx == 0),
                    show_xtick_labels=(comp_idx == len(bar_axes) - 1),
                    y_step=(Y_STEP_OVERRIDES.get((group_name, comp_idx))
                            if safe_name == "Air_15bp" else None))
            bar_axes[0].set_title(
                f"η @ {MARKER_I / 1000:.1f} A cm$^{{-2}}$",
                fontsize=fs, fontweight="bold", pad=title_pad)
        else:
            # Absolute waterfall: per-component shared scale across both durations
            comp_ymaxes = [0.0] * 4
            for dl, _ in DURS:
                for s in sample_list:
                    vals = bar_data[dl].get(s)
                    if vals is not None:
                        for ci, v in enumerate(vals):
                            comp_ymaxes[ci] = max(comp_ymaxes[ci], v)
            comp_ymaxes = [math.ceil(m * 20) / 20 + 0.05 for m in comp_ymaxes]
            for comp_idx, ax_bar in enumerate(bar_axes):
                draw_component_bar_panel(
                    ax_bar, bar_data, comp_idx, sample_list,
                    y_max=comp_ymaxes[comp_idx], fs=fs,
                    show_legend=(comp_idx == 0),
                    show_xtick_labels=(comp_idx == len(bar_axes) - 1))
            bar_axes[0].set_title(
                f"η @ {MARKER_I / 1000:.1f} A cm$^{{-2}}$",
                fontsize=fs, fontweight="bold", pad=title_pad)
        # Place bar column y-labels at a consistent figure-fraction x.
        # Using fig.text at (pos.x0 - _BAR_LABEL_X_OFFSET) is independent of
        # tick-label width, so spacing never breaks when decimal places change.
        x_lbl = bar_axes[0].get_position().x0 - _BAR_LABEL_X_OFFSET
        for comp_idx, ax_bar in enumerate(bar_axes):
            pos = ax_bar.get_position()
            y_c = (pos.y0 + pos.y1) / 2
            fig.text(x_lbl, y_c, f"{_COMP_NAMES[comp_idx]} (V)",
                     ha="center", va="center", rotation=90, fontsize=fs)
    # ─────────────────────────────────────────────────────────────────────────

    fname = f"Overpot_dur_{safe_name}_{group_name}_300mA.png"
    for d in (OUT_DIR, OUT_DIR_260603):
        out = os.path.join(d, fname)
        fig.savefig(out, dpi=300, bbox_inches="tight")
        print(f"Saved: {out}")
    plt.close(fig)


for cond_folder, safe_name, _, is_15bp in CONDITIONS:
    samp_rows = safe_name in ("Air_0bp", "Air_15bp")
    delta     = (safe_name == "Air_15bp")
    build_dur_grid(MAIN_SAMPLES,  cond_folder, safe_name, is_15bp, "Main",
                   samples_as_rows=samp_rows, use_delta=delta)
    build_dur_grid(OTHER_SAMPLES, cond_folder, safe_name, is_15bp, "Other",
                   samples_as_rows=samp_rows, use_delta=delta)


# ── Matched-yscale variant: all bar subplots share the kinetic y-scale ────
FINAL_PLOTS_260603 = os.path.join(OUT_DIR_260603, "Final Plots")
os.makedirs(FINAL_PLOTS_260603, exist_ok=True)


def _draw_inset_grouped_bar(ax_parent, bar_data, comp_idx, sample_list,
                             y_max_inset, fs_ins=13):
    """Zoomed grouped-bar inset in the upper ~50% of ax_parent (natural scale)."""
    ax_ins = ax_parent.inset_axes([0.20, 0.46, 0.72, 0.50])
    n = len(sample_list)
    x = np.arange(n, dtype=float)
    w = 0.35
    bol_vals = np.array([
        bar_data["BOL"][s][comp_idx] if bar_data["BOL"].get(s) is not None else 0.0
        for s in sample_list])
    k75_vals = np.array([
        bar_data["75K"][s][comp_idx] if bar_data["75K"].get(s) is not None else 0.0
        for s in sample_list])
    ax_ins.bar(x - w / 2, bol_vals, width=w, facecolor="#cccccc",
               edgecolor="black", linewidth=0.5)
    ax_ins.bar(x + w / 2, k75_vals, width=w, facecolor="#555555",
               edgecolor="black", linewidth=0.5)
    ax_ins.set_xticks(x)
    ax_ins.set_xticklabels([])
    ax_ins.set_xlim(-0.6, n - 0.4)
    ax_ins.set_ylim(0, y_max_inset)
    if y_max_inset <= 0.015:
        ax_ins.yaxis.set_major_locator(plt.MultipleLocator(0.005))
        ax_ins.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.3f}"))
    elif y_max_inset <= 0.030:
        ax_ins.yaxis.set_major_locator(plt.MultipleLocator(0.01))
        ax_ins.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max_inset <= 0.08:
        ax_ins.yaxis.set_major_locator(plt.MultipleLocator(0.02))
        ax_ins.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    elif y_max_inset <= 0.20:
        ax_ins.yaxis.set_major_locator(plt.MultipleLocator(0.05))
        ax_ins.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2f}"))
    else:
        ax_ins.yaxis.set_major_locator(plt.MultipleLocator(0.10))
        ax_ins.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}"))
    ax_ins.tick_params(direction="out", top=False, right=False,
                       labelsize=fs_ins, width=0.8, length=3)
    ax_ins.patch.set_facecolor("white")
    for spine in ax_ins.spines.values():
        spine.set_linewidth(0.8)


def build_dur_grid_matched_yscale(sample_list, cond_folder, safe_name, group_name):
    """
    Samples-as-rows, 2-dur IV grid + bar column.  All four bar subplots share
    the kinetic overpotential y-scale.  Non-kinetic components that are less
    than 40% of that scale receive a zoomed inset for inter-sample comparison.
    Output → 260603.../Final Plots/*_matched.png.
    """
    nrows     = len(sample_list)
    ncols     = len(DURS)
    fs        = 18
    ymax_iv   = 1.18
    title_pad = fs * 0.67

    FIG_W_IV  = 5.0
    FIG_W_BAR = 7.0
    total_w   = FIG_W_IV * ncols + FIG_W_BAR
    _s        = (FIG_W_IV * ncols) / total_w
    ROW_LABEL_X = 0.04 * _s
    YLABEL_X    = 0.11 * _s
    LEFT        = 0.20 * _s

    fig = plt.figure(figsize=(total_w, 3.5 * nrows), dpi=300)
    gs  = fig.add_gridspec(nrows, ncols + 1,
                            width_ratios=[FIG_W_IV, FIG_W_IV, FIG_W_BAR])
    axes     = np.array([[fig.add_subplot(gs[r, c]) for c in range(ncols)]
                          for r in range(nrows)])
    bar_axes = [fig.add_subplot(gs[r, ncols]) for r in range(nrows)]
    bar_data = {dl: {} for dl, _ in DURS}

    fig.subplots_adjust(left=LEFT, hspace=0.25, wspace=_BAR_WSPACE)
    iv_xc = (axes[0, 0].get_position().x0 + axes[0, ncols - 1].get_position().x1) / 2
    fig.supxlabel(r"Current Density (Acm$^{-2}$)", fontsize=fs, y=0.06, x=iv_xc)
    fig.supylabel("Cell Voltage (V)", fontsize=fs, x=YLABEL_X)

    # ── IV panels ──────────────────────────────────────────────────────────
    for row_idx, samp in enumerate(sample_list):
        for col_idx, (dur_label, dur_folder) in enumerate(DURS):
            ax         = axes[row_idx, col_idx]
            is_topleft = (row_idx == 0 and col_idx == 0)
            stem = FILE_STEMS.get((cond_folder, samp))
            fp   = find_file(cond_folder, dur_folder, stem) if stem else None
            if fp is None:
                ax.text(0.5, 0.5, "no data", transform=ax.transAxes,
                        ha="center", va="center", fontsize=fs)
                ax.set_xlim(0, 3.5); ax.set_ylim(0.2, ymax_iv)
                bar_data[dur_label][samp] = None
            else:
                try:
                    I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                    if samp in EIS_RPRO:
                        R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                    erev = calc_erev(T_K, P_H2, P_O2)
                    draw_panel(ax, I_mA, V, R_ohm, R_pro, erev, intercept, slope,
                               fs=fs, lw=1.6, show_legend=is_topleft, ymax=ymax_iv,
                               legend_frameon=(group_name == "Main"))
                    bar_data[dur_label][samp] = calc_bar_values(
                        I_mA, V, R_ohm, R_pro, erev, intercept, slope)
                except Exception as e:
                    print(f"  ERROR {samp}/{dur_label}/{cond_folder}: {e}")
                    ax.text(0.5, 0.5, "error", transform=ax.transAxes,
                            ha="center", va="center", fontsize=fs)
                    bar_data[dur_label][samp] = None
            if row_idx == 0:
                ax.set_title(dur_label, fontsize=fs, fontweight="bold", pad=title_pad)
            if col_idx == 0:
                pos   = axes[row_idx, 0].get_position()
                y_fig = (pos.y0 + pos.y1) / 2
                fig.text(ROW_LABEL_X, y_fig, LABELS[samp],
                         fontsize=fs, fontweight="bold",
                         ha="center", va="center", rotation=90, clip_on=False)

    # ── Bar column: shared kinetic y-scale, insets for small components ────
    comp_raw_max = [0.0] * 4
    for dl, _ in DURS:
        for s in sample_list:
            vals = bar_data[dl].get(s)
            if vals:
                for ci, v in enumerate(vals):
                    comp_raw_max[ci] = max(comp_raw_max[ci], v)
    comp_ymaxes_natural = [max(_nice_upper(m), 0.01) for m in comp_raw_max]
    shared_ymax = comp_ymaxes_natural[0]   # kinetic component sets the common scale

    bar_axes[0].set_title(
        f"η @ {MARKER_I / 1000:.1f} A cm$^{{-2}}$",
        fontsize=fs, fontweight="bold", pad=title_pad)

    for comp_idx, ax_bar in enumerate(bar_axes):
        draw_grouped_bar_panel(ax_bar, bar_data, comp_idx, sample_list,
                               y_max=shared_ymax, fs=fs,
                               show_legend=(comp_idx == 0),
                               show_xtick_labels=(comp_idx == len(bar_axes) - 1))
        nat_max = comp_ymaxes_natural[comp_idx]
        if comp_idx > 0 and nat_max < 0.4 * shared_ymax:
            _draw_inset_grouped_bar(ax_bar, bar_data, comp_idx, sample_list,
                                    nat_max, fs_ins=fs - 5)

    x_lbl = bar_axes[0].get_position().x0 - _BAR_LABEL_X_OFFSET
    for comp_idx, ax_bar in enumerate(bar_axes):
        pos = ax_bar.get_position()
        y_c = (pos.y0 + pos.y1) / 2
        fig.text(x_lbl, y_c, f"{_COMP_NAMES[comp_idx]} (V)",
                 ha="center", va="center", rotation=90, fontsize=fs)

    fname = f"Overpot_dur_{safe_name}_{group_name}_300mA_matched.png"
    out   = os.path.join(FINAL_PLOTS_260603, fname)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {out}")


build_dur_grid_matched_yscale(MAIN_SAMPLES,  "Air BP", "Air_15bp", "Main")
build_dur_grid_matched_yscale(OTHER_SAMPLES, "Air BP", "Air_15bp", "Other")


# ── Standalone overpotential-component bar column (Air 1.5bp, @ 1.0 A cm⁻²) ─
def build_eta_bar_only_column(sample_list, target_mA, group_name, out_filename,
                               y_overrides=None, target_V=None):
    """Overpotential-component bar column for the Air BP condition.
    If target_V is given, components are evaluated at the current density where
    the I-V curve crosses that voltage; otherwise at the fixed target_mA."""
    fs = 18
    bar_data = {dl: {} for dl, _ in DURS}
    for samp in sample_list:
        for dur_label, dur_folder in DURS:
            stem = FILE_STEMS.get(("Air BP", samp))
            fp   = find_file("Air BP", dur_folder, stem) if stem else None
            if fp is None:
                bar_data[dur_label][samp] = None
                continue
            try:
                I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope = read_excel(fp)
                if samp in EIS_RPRO:
                    R_pro = np.full(len(R_pro), EIS_RPRO[samp])
                erev = calc_erev(T_K, P_H2, P_O2)
                if target_V is not None:
                    i_eval = current_at_voltage(I_mA, V, target_V)
                    if np.isnan(i_eval):
                        bar_data[dur_label][samp] = None
                        continue
                else:
                    i_eval = target_mA
                bar_data[dur_label][samp] = calc_bar_values(
                    I_mA, V, R_ohm, R_pro, erev, intercept, slope, i=i_eval)
            except Exception as e:
                print(f"  ERROR {samp}/{dur_label}/Air BP: {e}")
                bar_data[dur_label][samp] = None

    comp_ymaxes = [0.0] * 4
    for dl, _ in DURS:
        for s in sample_list:
            vals = bar_data[dl].get(s)
            if vals is not None:
                for ci, v in enumerate(vals):
                    comp_ymaxes[ci] = max(comp_ymaxes[ci], v)
    comp_ymaxes = [max(_nice_upper(m), 0.01) for m in comp_ymaxes]
    y_steps = [None] * 4
    y_mins  = [0.0] * 4
    if y_overrides:
        for comp_idx, ov in y_overrides.items():
            if len(ov) == 3:                       # (y_min, y_max, y_step)
                y_mins[comp_idx], comp_ymaxes[comp_idx], y_steps[comp_idx] = ov
            else:                                  # (y_max, y_step)
                comp_ymaxes[comp_idx], y_steps[comp_idx] = ov

    fig = plt.figure(figsize=(4.5, 3.2 * 4), constrained_layout=True)
    gs  = fig.add_gridspec(4, 1)
    axes = [fig.add_subplot(gs[r, 0]) for r in range(4)]
    for comp_idx, ax in enumerate(axes):
        draw_grouped_bar_panel(
            ax, bar_data, comp_idx, sample_list,
            y_max=comp_ymaxes[comp_idx], fs=fs,
            show_legend=(comp_idx == 0),
            show_xtick_labels=(comp_idx == len(axes) - 1),
            y_step=y_steps[comp_idx], y_min=y_mins[comp_idx])

    out = os.path.join(FINAL_PLOTS_260603, out_filename)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    print(f"Saved: {out}")
    plt.close(fig)


# 1.0A figures now evaluated at 0.4 V; 0.5A figures at 0.7 V (current value unused)
# Row 1 (Kinetic): y-axis 0.4–0.6 V, 0.05 per tick
_KIN_YOV = {0: (0.40, 0.60, 0.05)}
build_eta_bar_only_column(MAIN_SAMPLES,  1000.0, "Main",  "Bar_Eta_1p0_Main.png",  target_V=0.40, y_overrides=_KIN_YOV)
build_eta_bar_only_column(OTHER_SAMPLES, 1000.0, "Other", "Bar_Eta_1p0_Other.png", target_V=0.40, y_overrides=_KIN_YOV)
build_eta_bar_only_column(MAIN_SAMPLES,  500.0,  "Main",  "Bar_Eta_0p5_Main.png",  target_V=0.70, y_overrides=_KIN_YOV)
build_eta_bar_only_column(OTHER_SAMPLES, 500.0,  "Other", "Bar_Eta_0p5_Other.png", target_V=0.70, y_overrides=_KIN_YOV)

print("Done.")
