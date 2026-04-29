"""
Batch overpotential breakdown plots.
Auto-discovers every .xlsx in Overpotnital/**/Edited/, reads all parameters
directly from each workbook (T, P_H2, P_O2, Tafel intercept/slope),
and saves one PNG per file into Overpotential Graphs/{condition}/{durability}/.
"""

import glob
import math
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE      = os.path.dirname(os.path.abspath(__file__))
OVER_DIR  = os.path.join(HERE, "Overpotnital")
OUT_BASE  = os.path.join(HERE, "Overpotential Graphs")

# ---------------------------------------------------------------------------
# Physics
# ---------------------------------------------------------------------------

def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    p_h2 = p_h2_kPa / 101.3
    p_o2 = p_o2_kPa / 101.3
    S = ((p_h2 ** 2) * p_o2) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23
            - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))

# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def read_excel(filepath):
    """Return (I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope)."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["Sheet1"]

    T_K       = ws["R2"].value
    P_H2      = ws["T2"].value
    P_O2      = ws["U2"].value
    intercept = ws["S4"].value
    slope     = ws["S5"].value

    # Validate params
    for name, val in [("T_K", T_K), ("P_H2", P_H2), ("P_O2", P_O2),
                      ("intercept", intercept), ("slope", slope)]:
        if not isinstance(val, (int, float)):
            raise ValueError(f"Cell {name} is not numeric: {val!r}")

    I_list, V_list, R_ohm_list, R_pro_list = [], [], [], []
    for row in ws.iter_rows(min_row=4, values_only=True):
        i_raw = row[0]
        v_raw = row[1]
        r_ohm = row[3]
        r_pro = row[4]
        if not isinstance(i_raw, (int, float)) or i_raw <= 0:
            continue
        if not isinstance(v_raw, (int, float)):
            continue
        I_list.append(i_raw * 200)          # mA cm⁻²  (area = 5 cm²)
        V_list.append(float(v_raw))
        R_ohm_list.append(float(r_ohm) if isinstance(r_ohm, (int, float)) else np.nan)
        R_pro_list.append(float(r_pro) if isinstance(r_pro, (int, float)) else np.nan)

    if len(I_list) < 5:
        raise ValueError("Too few data rows")

    return (np.array(I_list), np.array(V_list),
            np.array(R_ohm_list), np.array(R_pro_list),
            float(T_K), float(P_H2), float(P_O2),
            float(intercept), float(slope))

# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------

def make_plot(I_mA, V, R_ohm, R_pro, T_K, P_H2, P_O2, intercept, slope, out_path, ymax=1.2):
    FS  = 14
    LW  = 1.5
    BW  = 1.0
    DPI = 300

    # Reset state for each figure
    matplotlib.rcParams.update(matplotlib.rcParamsDefault)
    matplotlib.rcParams.update({
        "font.family":       "sans-serif",
        "font.size":         FS,
        "axes.labelsize":    FS,
        "axes.titlesize":    FS,
        "xtick.labelsize":   FS,
        "ytick.labelsize":   FS,
        "legend.fontsize":   FS,
        "axes.linewidth":    BW,
        "xtick.major.width": BW,
        "ytick.major.width": BW,
    })

    erev      = calc_erev(T_K, P_H2, P_O2)
    I_A_total = I_mA / 200
    I_Acm2    = I_mA / 1000
    ln_j      = np.log(I_Acm2)

    eta_kin  = -(intercept + slope * ln_j)
    eta_ohm  = R_ohm * I_A_total
    eta_pro  = R_pro * I_A_total
    eta_mass = (erev - V) - eta_kin - eta_ohm - eta_pro

    V_erev = np.full_like(I_mA, erev)
    V_kin  = erev - eta_kin
    V_ohm  = V_kin - eta_ohm
    V_pro  = V_ohm - eta_pro
    V_pro_clipped = np.maximum(V_pro, V)

    c_kin  = "#FFA07A"
    c_ohm  = "#90C090"
    c_pro  = "#9090D8"
    c_mass = "#F0E860"

    fig, ax = plt.subplots(figsize=(5.5, 4.5))

    ax.fill_between(I_mA, V_kin, V_erev,        color=c_kin,  alpha=0.85, label="Kinetic")
    ax.fill_between(I_mA, V_ohm, V_kin,          color=c_ohm,  alpha=0.85, label="Ohmic")
    ax.fill_between(I_mA, V_pro, V_ohm,          color=c_pro,  alpha=0.85, label="Proton")
    ax.fill_between(I_mA, V,     V_pro_clipped,  color=c_mass, alpha=0.85, label="Mass Transport")

    ax.plot(I_mA, V, color="black", linewidth=LW, zorder=5)

    # Red dots at 500 mA cm⁻²
    i_mark      = 500.0
    v_kin_mark  = float(np.interp(i_mark, I_mA, V_kin))
    v_meas_mark = float(np.interp(i_mark, I_mA, V))

    dot_kw = dict(color="red", marker="o", markersize=8, linestyle="none", zorder=10)
    ax.plot(i_mark, v_kin_mark,  **dot_kw)
    ax.plot(i_mark, v_meas_mark, **dot_kw)
    lbl_kw = dict(color="red", fontsize=FS, va="center", ha="left")
    ax.text(i_mark + 50, v_kin_mark,  f"{v_kin_mark:.2f}",  **lbl_kw)
    ax.text(i_mark + 50, v_meas_mark, f"{v_meas_mark:.2f}", **lbl_kw)

    ax.axhline(erev, color="grey", linewidth=0.8, linestyle="--", alpha=0.6)

    ax.set_xlim(0, 3500)
    ax.set_ylim(0.2, ymax)
    ax.set_xlabel(r"Current Density (mAcm$^{-2}$)", fontsize=FS)
    ax.set_ylabel("Cell Voltage (V)", fontsize=FS)
    ax.tick_params(direction="out", top=False, right=False,
                   labelsize=FS, width=BW, length=5)

    leg = ax.legend(loc="upper right", fontsize=FS, frameon=True,
                    facecolor="white", edgecolor="black")
    leg.get_frame().set_linewidth(1.5)

    plt.tight_layout()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    plt.savefig(out_path, dpi=DPI, bbox_inches="tight")
    plt.close(fig)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    pattern = os.path.join(OVER_DIR, "**", "Edited", "*.xlsx")
    files   = sorted(glob.glob(pattern, recursive=True))
    print(f"Found {len(files)} files\n")

    ok, skipped, errors = 0, 0, 0

    for fp in files:
        # Skip Excel lock files created when a workbook is open
        if os.path.basename(fp).startswith("~$"):
            continue

        # Parse condition and durability from path
        rel   = os.path.relpath(fp, OVER_DIR)          # e.g. O2 BP\BOL\Edited\CN BM...xlsx
        parts = rel.split(os.sep)
        condition  = parts[0]                           # "O2 BP", "Air", etc.
        durability = parts[1].upper()                   # normalise 75k → 75K
        stem       = os.path.splitext(parts[-1])[0]    # filename without .xlsx

        out_path = os.path.join(OUT_BASE, condition, durability,
                                stem + "_overpot_breakdown.png")
        ymax = 1.18 if " BP" not in condition else 1.2

        try:
            data = read_excel(fp)
            make_plot(*data, out_path=out_path, ymax=ymax)
            print(f"  OK  {condition}/{durability}/{stem}")
            ok += 1
        except Exception as e:
            print(f"  SKIP {condition}/{durability}/{stem}  - {e}")
            if "Too few" in str(e) or "not numeric" in str(e):
                skipped += 1
            else:
                errors += 1

    print(f"\nDone: {ok} saved, {skipped} skipped (no data), {errors} errors")


if __name__ == "__main__":
    main()
