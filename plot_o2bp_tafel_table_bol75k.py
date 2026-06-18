"""
Regenerates, from the new O2 BP reproducibility data (BOL/75K only):
  - Table_O2_15bp_{Main,Other}_alt.png   (Tafel slope/intercept table)
  - Tafel_combined_BM_Polyol_mix.png     (CN-BM, KB-BM, VC-Polyol, AB-Polyol)
  - Tafel_combined_Polyol_BM_rest.png    (CN-Polyol, KB-Polyol, VC-BM, AB-BM)

Tafel intercept/slope = S4/S5, used directly (no refit).
Scatter points = columns J (Ln(i)) / K (Total Opot); if a file has those
columns empty, J/K are derived from I (col A) / V (col B) and Erev
(from R2/T2/U2, same formula used elsewhere in this project).

Output: 260603 New data set (reproducibility)/Final Plots/
"""

import os, sys, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "260603 New data set (reproducibility)",
                         "Overpotential DATA raw alt", "O2 BP")
OUT_DIR  = os.path.join(BASE, "260603 New data set (reproducibility)", "Final Plots")

MAIN_SAMPLES  = ["CN BM", "KB BM", "VC Polyol", "AB Polyol"]
OTHER_SAMPLES = ["CN Polyol", "KB Polyol", "VC BM", "AB BM"]

LABELS = {
    "CN BM":     "CN-BM",    "KB BM":     "KB-BM",
    "VC Polyol": "VC-Polyol","AB Polyol": "AB-Polyol",
    "CN Polyol": "CN-Polyol","KB Polyol": "KB-Polyol",
    "VC BM":     "VC-BM",   "AB BM":     "AB-BM",
}

DURS = ["BOL", "30K", "75K"]

DUR_META = {
    "BOL": dict(ls="-",  marker="s", color="#1565C0"),   # blue
    "30K": dict(ls="-.", marker="o", color="#2E7D32"),   # green
    "75K": dict(ls="--", marker="^", color="#C62828"),   # red
}

# VC Polyol 30K: stored S5 (-0.4542) is a decimal-point entry error.
# Refitting the same window used for BOL/75K (S4 matches exactly,
# -0.54485) gives S5 = -0.04542, consistent with the BOL/75K trend.
TAFEL_OVERRIDES = {
    ("VC Polyol", "30K"): {"S5": -0.04542},
}

FS = 22
plt.rcParams.update({
    "font.family":     "Arial",
    "font.size":       FS,
    "axes.linewidth":  1.0,
    "xtick.direction": "out", "ytick.direction": "out",
    "xtick.top":       False, "ytick.right":     False,
})


def calc_erev(T_K, p_h2_kPa, p_o2_kPa):
    ph = p_h2_kPa / 101.3
    po = p_o2_kPa / 101.3
    S  = ((ph ** 2) * po) ** 2
    if S <= 0:
        return 1.23 - 0.0009 * (T_K - 298)
    return (1.23 - 0.0009 * (T_K - 298)
            + (2.303 * 8.314 * T_K / (4 * 96485)) * math.log10(S))


def find_file(samp, dur):
    stem   = f"{samp} 1.5bp O2"
    folder = os.path.join(DATA_DIR, samp, dur)
    for sub in ("", "Edited"):
        for sfx in ("_edited.xlsx", "_unchanged.xlsx", ".xlsx"):
            p = os.path.join(folder, sub, stem + sfx)
            if os.path.exists(p):
                return p
    return None


def read_tafel(fp):
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["Sheet1"]
    S4 = float(ws["S4"].value)
    S5 = float(ws["S5"].value)
    T_K, P_H2, P_O2 = ws["R2"].value, ws["T2"].value, ws["U2"].value
    erev = calc_erev(float(T_K), float(P_H2), float(P_O2))

    J, K = [], []
    need_fallback = False
    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, j, k = row[0], row[1], row[9], row[10]
        if not (isinstance(a, (int, float)) and isinstance(b, (int, float)) and a > 0):
            continue
        if isinstance(j, (int, float)) and isinstance(k, (int, float)):
            J.append(j); K.append(k)
        else:
            need_fallback = True

    if need_fallback or not J:
        J, K = [], []
        for row in ws.iter_rows(min_row=4, values_only=True):
            a, b = row[0], row[1]
            if isinstance(a, (int, float)) and isinstance(b, (int, float)) and a > 0:
                J.append(math.log(a * 0.2))
                K.append(b - erev)

    wb.close()
    return S4, S5, np.array(J), np.array(K)


# ── Read all data once ──────────────────────────────────────────────────────
cache = {}
for samp in MAIN_SAMPLES + OTHER_SAMPLES:
    for dur in DURS:
        fp = find_file(samp, dur)
        if fp is None:
            raise FileNotFoundError(f"No O2 BP file for {samp}/{dur}")
        S4, S5, J, K = read_tafel(fp)
        if (samp, dur) in TAFEL_OVERRIDES:
            ov = TAFEL_OVERRIDES[(samp, dur)]
            S4 = ov.get("S4", S4)
            S5 = ov.get("S5", S5)
        cache[(samp, dur)] = (S4, S5, J, K)
        print(f"  {samp:<12} {dur:<4}: {os.path.basename(fp)}")

all_J = np.concatenate([cache[k][2] for k in cache])
all_K = np.concatenate([cache[k][3] for k in cache])
XLIM = (-5, 1)
YLIM = (math.floor(all_K.min() * 5) / 5 - 0.05, math.ceil(all_K.max() * 5) / 5 + 0.05)
print(f"Shared xlim={XLIM}, ylim={YLIM}")


# ── Table ────────────────────────────────────────────────────────────────────
HDR_BG  = "#1E3A5F"
HDR_FG  = "white"
ROW_A   = "#D9E4F0"
ROW_B   = "#FFFFFF"
SAMP_FG = "#1E3A5F"

COL_LABELS = ["Sample", "Dur.", "Tafel Slope", "Tafel Intercept"]
COL_W      = [0.14, 0.07, 0.22, 0.22]


def build_table(sample_list, group_name):
    rows_data = []
    for samp in sample_list:
        for i, dur in enumerate(DURS):
            S4, S5, _, _ = cache[(samp, dur)]
            samp_label = LABELS[samp] if i == 0 else ""
            rows_data.append((samp_label, dur, S5, S4))  # slope=S5, intercept=S4

    cell_text = [[lbl, dur, f"{slope:.4f}", f"{intercept:.3f}"]
                  for lbl, dur, slope, intercept in rows_data]

    n_data_rows = len(cell_text)
    n_cols      = len(COL_LABELS)

    data_row_h_in = 0.34
    hdr_row_h_in  = 0.44
    fig_h = n_data_rows * data_row_h_in + hdr_row_h_in + 0.70
    fig_w = 7.0

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    tbl = ax.table(
        cellText=cell_text,
        colLabels=COL_LABELS,
        loc="upper center",
        cellLoc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9.5)

    hdr_h_frac  = hdr_row_h_in  / fig_h
    data_h_frac = data_row_h_in / fig_h

    for j in range(n_cols):
        tbl[0, j].set_height(hdr_h_frac)
    for i in range(1, n_data_rows + 1):
        for j in range(n_cols):
            tbl[i, j].set_height(data_h_frac)

    total_w = sum(COL_W)
    for j, w in enumerate(COL_W):
        for i in range(n_data_rows + 1):
            tbl[i, j].set_width(w / total_w)

    # Header styling
    for j in range(n_cols):
        cell = tbl[0, j]
        cell.set_facecolor(HDR_BG)
        cell.set_text_props(color=HDR_FG, fontweight="bold")
        cell.set_edgecolor("white")

    # Data row styling (2 rows per sample: BOL, 75K)
    for i, (samp_label, dur, *_) in enumerate(rows_data):
        tbl_row   = i + 1
        group_idx = i // len(DURS)
        bg = ROW_A if group_idx % 2 == 0 else ROW_B

        for j in range(n_cols):
            cell = tbl[tbl_row, j]
            cell.set_facecolor(bg)
            cell.set_edgecolor("#AAAAAA")

        if samp_label:
            tbl[tbl_row, 0].set_text_props(color=SAMP_FG, fontweight="bold")

        if i % len(DURS) == 0 and i > 0:
            for j in range(n_cols):
                tbl[tbl_row, j].set_edgecolor("#555555")

    fig.text(0.5, 0.98,
             rf"O$_2$ 1.5 bar BP  —  {group_name} samples  (BOL / 30K / 75K)",
             ha="center", va="top",
             fontsize=12, fontweight="bold", color=HDR_BG)

    out = os.path.join(OUT_DIR, f"Table_O2_15bp_{group_name}_alt.png")
    fig.savefig(out, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out}")


# ── Tafel combined ──────────────────────────────────────────────────────────
def build_tafel_combined(sample_list, out_name):
    n = len(sample_list)
    fig = plt.figure(figsize=(4.5 * n, 4.5), constrained_layout=True)
    gs  = fig.add_gridspec(1, n)
    axes = [fig.add_subplot(gs[0, i]) for i in range(n)]

    x_line = np.linspace(XLIM[0], XLIM[1], 200)

    for i, samp in enumerate(sample_list):
        ax = axes[i]
        for dur in DURS:
            S4, S5, J, K = cache[(samp, dur)]
            meta = DUR_META[dur]
            ax.plot(J, K, marker=meta["marker"], color=meta["color"], ms=6,
                    linestyle="none", markeredgewidth=0, zorder=3)
            ax.plot(x_line, S4 + S5 * x_line, color=meta["color"],
                    ls=meta["ls"], lw=1.8, zorder=2)

        ax.set_xlim(*XLIM)
        ax.set_ylim(*YLIM)
        ax.xaxis.set_major_locator(plt.MultipleLocator(2))
        ax.yaxis.set_major_locator(plt.MultipleLocator(0.2))
        ax.tick_params(direction="out", top=False, right=False,
                        labelsize=FS, width=1.0, length=5)
        for spine in ax.spines.values():
            spine.set_linewidth(1.0)
        ax.set_title(LABELS[samp], fontsize=FS, fontweight="bold")

        if i > 0:
            ax.tick_params(labelleft=False)

        if i == 0:
            handles = [mlines.Line2D([], [], color=DUR_META[d]["color"],
                                      marker=DUR_META[d]["marker"], ls=DUR_META[d]["ls"],
                                      ms=8, label=d) for d in DURS]
            ax.legend(handles=handles, loc="lower left", fontsize=FS - 4,
                      frameon=False, handlelength=1.5, borderpad=0.3, labelspacing=0.3)

    fig.supxlabel(r"ln|j|  (j in A cm$^{-2}$)", fontsize=FS)
    fig.supylabel(r"Total Overpotential ($\eta$)", fontsize=FS)

    out = os.path.join(OUT_DIR, out_name)
    fig.savefig(out, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {out}")


build_table(MAIN_SAMPLES,  "Main")
build_table(OTHER_SAMPLES, "Other")
build_tafel_combined(MAIN_SAMPLES,  "Tafel_combined_BM_Polyol_mix.png")
build_tafel_combined(OTHER_SAMPLES, "Tafel_combined_Polyol_BM_rest.png")

print("Done.")
